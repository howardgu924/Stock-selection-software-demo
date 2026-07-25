import json
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from stock_picker.strategies.adaptive_trend_v1_3 import (
    AccountSnapshot, DataSnapshot, NetworkAccessPolicy, ResolvedDateRange,
    RunConfig, RunMode, RunStore, UniverseSnapshot, create_run,
    generate_run_report,
)


def setup_store(tmp_path):
    store=RunStore(tmp_path/"runs.sqlite3")
    days=(date(2025,1,2),)
    resolved=ResolvedDateRange(days[0],days[0],days[0],days[0],days[0],days,days)
    config=RunConfig(
        RunMode.BACKTEST,"V1.3.13","account","universe","data",resolved,320,
        "RAW_UNADJUSTED_V1",NetworkAccessPolicy.FORBID,str(tmp_path),"EMPTY",
        "2025-01-01T00:00:00+08:00","config-hash",
        git_commit_sha="a"*40,schema_version=3,
    )
    account=AccountSnapshot("account","p",RunMode.BACKTEST,Decimal("1000"),(),"fees","CNY","2025-01-01T00:00:00+08:00","account-hash")
    universe=UniverseSnapshot("universe",(),(),(),("000300.SH","000905.SH","000852.SH"),(),"universe-hash","2025-01-01T00:00:00+08:00")
    data=DataSnapshot(
        "data",("partition",),"RAW_UNADJUSTED_V1","2025-01-01T00:00:00+08:00",
        "data-hash",(("partition","partition-hash"),),("2025-01-02",),
        ("rule-v1",),("fee-v1",),"READY",
        (("partition","daily_bar","logical","provider","provider-v1","1d"),),
    )
    create_run(store,config,{"account":account,"universe":universe,"data":data},run_id="run")
    store.update_run_status("run","COMPLETED")
    return store


def test_report_has_all_files_and_sheets(tmp_path):
    root=generate_run_report(setup_store(tmp_path),"run",tmp_path)
    expected={"backtest_report.xlsx","run_manifest.json","run_config.json","audit_log.jsonl","data_readiness.json"}
    assert expected=={path.name for path in root.iterdir()}
    workbook=load_workbook(root/"backtest_report.xlsx")
    assert workbook.sheetnames==["运行摘要","每日权益","基准对比","成交明细","订单与失败","每日持仓","候选与评分","退出与Pending","冷却期","数据覆盖","异常与警告"]


def test_excel_format_contract(tmp_path):
    root=generate_run_report(setup_store(tmp_path),"run",tmp_path)
    workbook=load_workbook(root/"backtest_report.xlsx")
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes=="A2"
        assert max((dimension.width or 0 for dimension in sheet.column_dimensions.values()),default=0)<=50


def test_manifest_contains_hashes_and_no_secret(tmp_path):
    root=generate_run_report(setup_store(tmp_path),"run",tmp_path,manifest_context={"password":"no","git_sha":"abc"})
    manifest=json.loads((root/"run_manifest.json").read_text(encoding="utf-8"))
    assert "password" not in manifest
    assert all(len(item["sha256"])==64 for item in manifest["files"])
    assert manifest["git_commit_sha"]=="a"*40
    assert "git_sha" not in manifest


def test_report_is_rebuildable_from_store(tmp_path):
    store=setup_store(tmp_path)
    first=generate_run_report(store,"run",tmp_path)
    (first/"backtest_report.xlsx").unlink()
    second=generate_run_report(store,"run",tmp_path)
    assert (second/"backtest_report.xlsx").exists()
