from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from stock_picker.reporting.t1_thermostat_backtest_report import (
    EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS,
    build_t1_thermostat_backtest_report,
    default_t1_thermostat_backtest_report_filename,
    export_t1_thermostat_backtest_excel,
)
from stock_picker.strategies.thermostat_backtest import (
    RESULT_TABLE_COLUMNS,
    T1ThermostatBacktestResult,
)


EXPECTED_SHEETS = [
    "回测摘要",
    "回测说明",
    "参数与账户设置",
    "数据来源与股票池",
    "每日资产",
    "权益与回撤",
    "每日持仓",
    "每日触发计划",
    "订单明细",
    "成交明细",
    "失败订单",
    "取消订单",
    "pending明细",
    "趋势批次",
    "网格层级",
    "个股表现",
    "趋势策略表现",
    "网格策略表现",
    "市场状态表现",
    "数据质量",
    "公司行为影响",
]


def _result() -> T1ThermostatBacktestResult:
    frames = {
        name: pd.DataFrame(columns=columns)
        for name, columns in RESULT_TABLE_COLUMNS.items()
    }
    frames["summary"] = pd.DataFrame(
        [{
            "initial_asset": 100_000.0,
            "final_asset": 105_000.0,
            "total_return": 0.05,
            "annualized_return": 0.12,
            "failed_order_count": 1,
            "pending_order_count": 1,
            "trading_day_count": 2,
        }]
    )
    frames["daily_assets"] = pd.DataFrame(
        [{
            "date": "2026-01-05",
            "cash": 89_000.0,
            "position_value": 11_000.0,
            "total_asset": 100_000.0,
            "cash_ratio": 0.89,
            "position_ratio": 0.11,
            "realized_pnl": 0.0,
            "unrealized_pnl": 100.0,
            "precision": "daily_approximate",
            "precision_disclosure": "daily-bar approximation",
            "approximate_intraday_sequence": True,
        }]
    )
    frames["daily_trigger_plans"] = pd.DataFrame(
        [{
            "date": "2026-01-05",
            "symbol": "600001.SH",
            "stock_mode": "trend",
            "plan_trace_id": "plan-1",
            "market_position_discount": 0.8,
            "target_position_pct": 0.2,
            "grid_layer_spacing_pct": 0.025,
            "precision": "daily_approximate",
            "precision_disclosure": "daily-bar approximation",
            "approximate_intraday_sequence": True,
        }]
    )
    order = {
        "order_id": "order-1",
        "trade_date": "2026-01-05",
        "symbol": "600001.SH",
        "mode": "trend",
        "family": "trend",
        "trigger_type": "trend_buy",
        "side": "buy",
        "status": "filled",
        "trigger_price": 10.0,
        "base_price": 10.1,
        "execution_price": 10.11,
        "intended_shares": 100,
        "actual_shares": 100,
        "gross_amount": 1011.0,
        "commission": 5.0,
        "stamp_tax": 0.0,
        "slippage_cost": 1.0,
        "total_cost": 1016.0,
        "cash_before": 100_000.0,
        "cash_after": 98_984.0,
        "position_before": 0,
        "position_after": 100,
        "approximate_intraday_sequence": True,
        "plan_trace_id": "plan-1",
        "candidate_trace_id": "candidate-1",
    }
    failed = {
        **order,
        "order_id": "order-failed",
        "status": "failed",
        "actual_shares": 0,
        "failure_reason": "limit_up_buy_failed",
        "plan_trace_id": "plan-failed",
        "candidate_trace_id": "candidate-failed",
    }
    cancelled = {
        **order,
        "order_id": "order-cancelled",
        "status": "cancelled",
        "actual_shares": 0,
        "failure_reason": "unfilled_cancelled",
        "plan_trace_id": "plan-cancelled",
        "candidate_trace_id": "candidate-cancelled",
    }
    frames["lifecycle_orders"] = pd.DataFrame([order, failed, cancelled])
    frames["fills"] = pd.DataFrame([order])
    frames["failed_cancelled_orders"] = pd.DataFrame([failed, cancelled])
    frames["pending_history"] = pd.DataFrame(
        [{
            "date": "2026-01-06",
            "symbol": "600001.SH",
            "episode_id": "episode-1",
            "event_type": "pending_retry",
            "is_terminal": False,
            "level": "pending_exit",
            "family": "trend",
            "owner_id": "batch-1",
            "remaining_shares": 100,
            "requested_shares": 100,
            "pending_since": "2026-01-05",
            "duration_days": 1,
            "attempt_count": 1,
            "last_attempt_date": "2026-01-06",
            "last_failure": "open_at_limit_down",
            "source_order_id": "order-failed",
            "plan_trace_id": "plan-failed",
        }]
    )
    frames["parameters"] = pd.DataFrame(
        [
            {"parameter_name": "commission_rate", "parameter_value": 0.0003, "parameter_source": "account_setting"},
            {"parameter_name": "minimum_commission", "parameter_value": 5.0, "parameter_source": "account_setting"},
            {"parameter_name": "stamp_tax_rate", "parameter_value": 0.001, "parameter_source": "account_setting"},
            {"parameter_name": "slippage_pct", "parameter_value": 0.001, "parameter_source": "account_setting"},
            {"parameter_name": "buy_lot_size", "parameter_value": 100, "parameter_source": "system_default"},
            {"parameter_name": "force_final_liquidation", "parameter_value": False, "parameter_source": "system_default"},
        ]
    )
    frames["stock_pool_metadata"] = pd.DataFrame(
        [
            {"metadata_key": "pool_type", "metadata_value": "market_range"},
            {"metadata_key": "membership", "metadata_value": "static current snapshot"},
            {"metadata_key": "generation_method", "metadata_value": "current market range"},
            {"metadata_key": "look_ahead_selection_warning", "metadata_value": "look-ahead selection bias"},
            {"metadata_key": "survivor_bias_warning", "metadata_value": "survivor bias"},
        ]
    )
    frames["closed_trade_cycles"] = pd.DataFrame(
        [{
            "cycle_id": "cycle-1",
            "symbol": "600001.SH",
            "family": "trend",
            "owner_id": "batch-1",
            "buy_order_id": "order-1",
            "sell_order_id": "order-2",
            "buy_date": "2026-01-05",
            "sell_date": "2026-01-08",
            "shares": 100,
            "buy_price": 10.11,
            "sell_price": 10.8,
            "buy_fees": 5.0,
            "sell_fees": 6.08,
            "gross_pnl": 69.0,
            "net_pnl": 57.92,
            "return_pct": 0.057,
            "holding_days": 3,
            "is_win": True,
        }]
    )
    return T1ThermostatBacktestResult(**frames)


def test_builds_exact_ordered_contract_without_mutating_raw_result() -> None:
    raw = _result()
    before = {name: frame.copy(deep=True) for name, frame in raw.tables.items()}

    report = build_t1_thermostat_backtest_report(
        raw, generated_at=datetime(2026, 7, 13, 9, 8, 7)
    )

    assert EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS == EXPECTED_SHEETS
    assert list(report.tables) == EXPECTED_SHEETS
    assert "holding_advice" not in "|".join(report.tables)
    assert report.tables["回测摘要"].loc[0, "期末资产"] == 105_000.0
    assert report.tables["回测摘要"].loc[0, "失败订单数"] == 1
    assert report.tables["回测摘要"].loc[0, "pending记录数"] == 1
    for name, frame in raw.tables.items():
        pd.testing.assert_frame_equal(frame, before[name])


def test_report_export_never_reads_legacy_advice_tables(tmp_path) -> None:
    class PoisonedResult(T1ThermostatBacktestResult):
        def __getattribute__(self, name):
            if name in {
                "_deprecated_signal_rows", "holding_advice", "new_candidates",
                "grid_advice", "trend_advice",
            }:
                raise AssertionError(f"legacy advice accessed: {name}")
            return super().__getattribute__(name)

    base = _result()
    raw = PoisonedResult(**{name: frame.copy(deep=True) for name, frame in base.tables.items()})

    report = build_t1_thermostat_backtest_report(raw)
    output = export_t1_thermostat_backtest_excel(report, tmp_path / "poison-safe.xlsx")

    assert output.exists()


def test_report_discloses_precision_fills_account_cache_and_pool_biases() -> None:
    report = build_t1_thermostat_backtest_report(_result())
    explanation = "\n".join(map(str, report.tables["回测说明"].to_numpy().ravel()))
    parameters = "\n".join(map(str, report.tables["参数与账户设置"].to_numpy().ravel()))
    pool = "\n".join(map(str, report.tables["数据来源与股票池"].to_numpy().ravel()))

    for disclosure in (
        "回测精度：日线近似",
        "分钟线：未使用",
        "盘中触发时间：无法准确识别",
        "同日多触发：使用保守顺序处理",
        "买入成交价",
        "卖出成交价",
        "涨停",
        "跌停",
        "停牌",
        "T+1",
    ):
        assert disclosure in explanation
    for setting in (
        "commission_rate",
        "minimum_commission",
        "stamp_tax_rate",
        "slippage_pct",
        "buy_lot_size",
        "trend_symbol_base_max",
        "trend_total_base_max",
        "grid_symbol_base_max",
        "grid_total_hard_max",
        "account_total_max",
        "趋势分批比例",
        "网格层数",
        "网格间距",
        "force_final_liquidation",
        "买入成交价公式",
        "卖出成交价公式",
    ):
        assert setting in parameters
    for disclosure in (
        "数据源",
        "前复权指标流",
        "不复权执行流",
        "static current snapshot",
        "current market range",
        "look-ahead selection bias",
        "survivor bias",
    ):
        assert disclosure in pool


def test_export_retains_trace_ids_completed_cycles_and_readable_empty_sheets(tmp_path: Path) -> None:
    report = build_t1_thermostat_backtest_report(_result())
    output = export_t1_thermostat_backtest_excel(report, tmp_path / "report.xlsx")
    book = load_workbook(output, data_only=False)

    assert book.sheetnames == EXPECTED_SHEETS
    for sheet_name in ("每日持仓", "趋势批次", "网格层级", "数据质量", "公司行为影响"):
        assert book[sheet_name]["A1"].value == "暂无数据"
    for sheet_name, trace in (
        ("订单明细", "candidate-1"),
        ("失败订单", "candidate-failed"),
        ("取消订单", "candidate-cancelled"),
        ("pending明细", "plan-failed"),
    ):
        values = [cell.value for row in book[sheet_name].iter_rows() for cell in row]
        assert trace in values
    completed_values = [cell.value for row in book["成交明细"].iter_rows() for cell in row]
    assert "cycle-1" in completed_values
    assert "已完成交易周期" in completed_values


def test_export_applies_styles_width_caps_and_semantic_formats(tmp_path: Path) -> None:
    report = build_t1_thermostat_backtest_report(_result())
    output = export_t1_thermostat_backtest_excel(report, tmp_path / "styled.xlsx")
    book = load_workbook(output)
    assets = book["每日资产"]
    plans = book["每日触发计划"]
    orders = book["订单明细"]

    assert assets.freeze_panes == "A2"
    assert assets.auto_filter.ref == assets.dimensions
    assert all(cell.font.bold for cell in assets[1])
    assert all(cell.alignment.vertical == "top" for row in assets.iter_rows() for cell in row)
    assert all((dimension.width or 0) <= 48 for dimension in assets.column_dimensions.values())
    assert all((dimension.width or 0) <= 48 for dimension in plans.column_dimensions.values())

    def body_format(sheet, header: str) -> str:
        headers = [cell.value for cell in sheet[1]]
        return sheet.cell(row=2, column=headers.index(header) + 1).number_format

    assert body_format(assets, "日期") == "yyyy-mm-dd"
    assert body_format(assets, "现金") == "#,##0.00"
    assert body_format(assets, "现金比例") == "0.00%"
    assert body_format(plans, "股票代码") == "@"
    assert body_format(orders, "订单ID") == "@"
    assert body_format(orders, "实际股数") == "0"
    assert body_format(orders, "成交价格") == "0.000"
    long_text_headers = {"精度说明", "近似处理警告", "失败原因", "质量警告"}
    for sheet in (assets, plans, orders):
        for cell in sheet[1]:
            if cell.value in long_text_headers:
                assert sheet.cell(row=2, column=cell.column).alignment.wrap_text is True


def test_timestamped_filename_is_stable_and_second_specific() -> None:
    first = default_t1_thermostat_backtest_report_filename(
        datetime(2026, 7, 13, 9, 8, 7)
    )
    second = default_t1_thermostat_backtest_report_filename(
        datetime(2026, 7, 13, 9, 8, 8)
    )

    assert first == "t1_thermostat_backtest_20260713_090807.xlsx"
    assert second != first
    assert re.fullmatch(r"t1_thermostat_backtest_\d{8}_\d{6}\.xlsx", first)


def test_consecutive_default_reports_have_unique_second_filenames() -> None:
    first = build_t1_thermostat_backtest_report(_result())
    second = build_t1_thermostat_backtest_report(_result())

    assert (
        default_t1_thermostat_backtest_report_filename(first)
        != default_t1_thermostat_backtest_report_filename(second)
    )


def test_expired_untriggered_plan_keeps_order_and_plan_ids_without_fabricating_candidate() -> None:
    raw = _result()
    expired = raw.lifecycle_orders.iloc[0].copy()
    expired["order_id"] = "order-expired"
    expired["status"] = "expired"
    expired["plan_trace_id"] = "plan-expired"
    expired["candidate_trace_id"] = ""
    raw.failed_cancelled_orders = pd.concat(
        [raw.failed_cancelled_orders, pd.DataFrame([expired])],
        ignore_index=True,
    )
    report = build_t1_thermostat_backtest_report(raw)
    failed = report.tables["失败订单"]
    row = failed.loc[failed["订单ID"].eq("order-expired")].iloc[0]
    explanation = "\n".join(map(str, report.tables["回测说明"].to_numpy().ravel()))

    assert row["订单ID"] == "order-expired"
    assert row["计划追踪ID"] == "plan-expired"
    assert row["候选追踪ID"] == ""
    assert "未创建执行候选时，candidate_trace_id按设计留空" in explanation
