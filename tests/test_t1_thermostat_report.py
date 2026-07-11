from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from stock_picker.reporting.t1_thermostat_report import (
    EXPECTED_T1_THERMOSTAT_SHEETS,
    build_t1_thermostat_report,
    default_t1_thermostat_report_filename,
    export_t1_thermostat_excel,
)


def _trigger_plan() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "date": "2026-07-08",
                "symbol": "600001.SH",
                "name": "Trend",
                "stock_mode": "trend",
                "market_regime": "market_uptrend",
                "market_regime_normalized": "strong",
                "market_position_discount": 1.0,
                "target_position_pct": 0.2,
                "max_position_pct": 0.2,
                "total_shares": 100,
                "available_shares": 0,
                "today_bought_shares": 100,
                "pending_sell_level": "pending_exit",
                "boll_upper": 12.8,
                "boll_mid": 11.5,
                "boll_lower": 10.2,
                "atr20": 0.42,
                "trend_buy_trigger": 12.9,
                "trend_reduce_trigger": 11.5,
                "trend_exit_trigger": 10.0,
                "trend_batches": "40%,35%,25%",
                "grid_lower": "",
                "grid_mid": "",
                "grid_upper": "",
                "grid_max_layers": "",
                "grid_buy_levels": "",
                "grid_sell_levels": "",
                "grid_total_max_position_pct": 0.0,
                "trigger_status": "triggered",
                "filled_status": "pending",
                "failed_reason": "",
                "risk_note": "今日买入不可卖",
                "reason": "趋势计划",
            },
            {
                "date": "2026-07-08",
                "symbol": "600002.SH",
                "name": "Range",
                "stock_mode": "range",
                "market_regime": "market_range",
                "market_regime_normalized": "normal",
                "market_position_discount": 0.9,
                "target_position_pct": 0.135,
                "max_position_pct": 0.135,
                "total_shares": 0,
                "available_shares": 0,
                "today_bought_shares": 0,
                "pending_sell_level": "",
                "boll_upper": "",
                "boll_mid": "",
                "boll_lower": "",
                "atr20": "",
                "trend_buy_trigger": "",
                "trend_reduce_trigger": "",
                "trend_exit_trigger": "",
                "trend_batches": "",
                "grid_lower": 9.1,
                "grid_mid": 10.0,
                "grid_upper": 10.9,
                "grid_max_layers": 3,
                "grid_buy_levels": "9.65|9.30|9.10",
                "grid_sell_levels": "10.35|10.70|10.90",
                "grid_total_max_position_pct": 0.4,
                "trigger_status": "planned",
                "filled_status": "failed",
                "failed_reason": "limit_up_buy_failed",
                "risk_note": "",
                "reason": "网格计划",
            },
        ]
    )


def test_t1_thermostat_excel_contains_expected_sheets_and_formatting(tmp_path: Path) -> None:
    report = build_t1_thermostat_report(_trigger_plan(), pd.DataFrame([{"symbol": "600003.SH", "error": "missing history"}]))
    output = tmp_path / default_t1_thermostat_report_filename(report)

    export_t1_thermostat_excel(report, output)

    book = load_workbook(output)
    assert book.sheetnames == EXPECTED_T1_THERMOSTAT_SHEETS
    assert not {"Holding Advice", "New Buy Candidates", "Grid Advice", "Trend Advice", "holding_advice", "new_candidates", "grid_advice", "trend_advice"} & set(book.sheetnames)

    summary = book["个股模式摘要"]
    headers = [cell.value for cell in summary[1]]
    assert {"日期", "股票", "名称", "股票模式", "目标仓位", "单股仓位上限", "今日买入股数", "待卖级别"}.issubset(headers)
    assert summary.freeze_panes == "A2"
    assert summary["A1"].font.bold
    assert summary.column_dimensions["A"].width >= 10
    assert summary.cell(row=2, column=headers.index("目标仓位") + 1).number_format == "0.00%"

    trend = book["趋势触发计划"]
    trend_headers = [cell.value for cell in trend[1]]
    assert {"布林上轨", "ATR20", "趋势买入触发价", "趋势分批"}.issubset(trend_headers)
    assert trend.cell(row=2, column=trend_headers.index("布林上轨") + 1).number_format == "0.00"

    grid = book["网格触发计划"]
    grid_headers = [cell.value for cell in grid[1]]
    assert {"网格下沿", "最大网格层数", "网格买入层", "网格总仓位上限"}.issubset(grid_headers)
    assert grid.cell(row=2, column=grid_headers.index("网格买入层") + 1).alignment.wrap_text
    assert grid.cell(row=2, column=grid_headers.index("网格总仓位上限") + 1).number_format == "0.00%"


def test_t1_thermostat_excel_writes_readable_empty_sheets(tmp_path: Path) -> None:
    report = build_t1_thermostat_report(pd.DataFrame(), pd.DataFrame())
    output = tmp_path / default_t1_thermostat_report_filename(report)

    export_t1_thermostat_excel(report, output)

    book = load_workbook(output)
    assert book.sheetnames == EXPECTED_T1_THERMOSTAT_SHEETS
    for sheet_name in EXPECTED_T1_THERMOSTAT_SHEETS:
        assert book[sheet_name]["A1"].value == "暂无数据"

