from __future__ import annotations

import json
import re
import threading
from pathlib import Path
from types import SimpleNamespace
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook

from examples import web_app
from stock_picker.reporting.t1_thermostat_report import build_t1_thermostat_report, export_t1_thermostat_excel
from stock_picker.reporting.t1_thermostat_backtest_report import (
    EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS,
)
from stock_picker.strategies.thermostat_backtest import RESULT_TABLE_COLUMNS, T1ThermostatBacktestResult
from stock_picker.user import ManualPortfolioStore, WatchlistStore


MOJIBAKE_MARKERS = ("锛", "涓", "浠", "绯", "鎭", "鍔", "瀹", "�")


def _assert_no_mojibake(text: str, *, context: str) -> None:
    found = [marker for marker in MOJIBAKE_MARKERS if marker in text]
    assert not found, f"{context} contains mojibake markers: {found}"


def _assert_no_raw_internal_fields(text: str, fields: tuple[str, ...]) -> None:
    leaked = [field for field in fields if field in text]
    assert not leaked, f"rendered output leaked internal fields: {leaked}"


def _history(symbol: str, closes: list[float]) -> pd.DataFrame:
    dates = pd.date_range("2026-04-01", periods=len(closes), freq="D")
    return pd.DataFrame(
        {
            "symbol": symbol,
            "date": dates.strftime("%Y-%m-%d"),
            "open": closes,
            "high": [value * 1.01 for value in closes],
            "low": [value * 0.99 for value in closes],
            "close": closes,
            "volume": [100000] * len(closes),
        }
    )


class FakeWebService:
    def __init__(self, histories: dict[str, pd.DataFrame] | None = None, quotes: pd.DataFrame | None = None) -> None:
        self.histories = histories or {}
        self.quotes = quotes if quotes is not None else pd.DataFrame(columns=["symbol", "name", "price"])

    def get_history(self, symbol: str, **kwargs) -> pd.DataFrame:
        return self.histories[symbol]

    def get_realtime_quotes(self, symbols=None) -> pd.DataFrame:
        if symbols:
            wanted = set(symbols)
            return self.quotes[self.quotes["symbol"].isin(wanted)].reset_index(drop=True)
        return self.quotes

    def get_market_snapshot(self, symbols=None) -> pd.DataFrame:
        return self.get_realtime_quotes(symbols)

    def get_stock_symbols(self, refresh: bool = False):
        return []

    def get_index_history(self, index_code: str, start_date: str, end_date: str, period: str = "daily") -> pd.DataFrame:
        return _history("000001.SH", [3000 + i * 10 for i in range(40)])


def _minimal_backtest_result() -> T1ThermostatBacktestResult:
    frames = {name: pd.DataFrame(columns=columns) for name, columns in RESULT_TABLE_COLUMNS.items()}
    frames["summary"] = pd.DataFrame([{"initial_asset": 100000.0, "final_asset": 100000.0, "total_return": 0.0}])
    return T1ThermostatBacktestResult(**frames)


def _sample_trigger_plan() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "date": "2026-07-08",
                "symbol": "600001.SH",
                "name": "Trend",
                "stock_mode": "trend",
                "market_regime": "market_uptrend",
                "market_regime_normalized": "strong",
                "market_position_discount": 1.0,
                "target_position_pct": 0.2,
                "max_position_pct": 0.2,
                "total_shares": 100,
                "available_shares": 0,
                "today_bought_shares": 100,
                "pending_sell_level": "pending_exit",
                "boll_upper": 12.8,
                "boll_mid": 11.5,
                "boll_lower": 10.2,
                "atr20": 0.42,
                "trend_buy_trigger": 12.9,
                "trend_reduce_trigger": 11.5,
                "trend_exit_trigger": 10.0,
                "trend_batches": "40%,35%,25%",
                "grid_lower": "",
                "grid_mid": "",
                "grid_upper": "",
                "grid_max_layers": "",
                "grid_buy_levels": "",
                "grid_sell_levels": "",
                "grid_total_max_position_pct": 0.0,
                "trigger_status": "triggered",
                "filled_status": "pending",
                "failed_reason": "",
                "risk_note": "今日买入不可卖",
                "reason": "趋势计划",
            },
            {
                "date": "2026-07-08",
                "symbol": "600002.SH",
                "name": "Range",
                "stock_mode": "range",
                "market_regime": "market_range",
                "market_regime_normalized": "normal",
                "market_position_discount": 0.9,
                "target_position_pct": 0.135,
                "max_position_pct": 0.135,
                "total_shares": 0,
                "available_shares": 0,
                "today_bought_shares": 0,
                "pending_sell_level": "",
                "grid_lower": 9.1,
                "grid_mid": 10.0,
                "grid_upper": 10.9,
                "grid_max_layers": 3,
                "grid_buy_levels": "9.65|9.30|9.10",
                "grid_sell_levels": "10.35|10.70|10.90",
                "grid_total_max_position_pct": 0.4,
                "trigger_status": "planned",
                "filled_status": "failed",
                "failed_reason": "limit_up_buy_failed",
                "risk_note": "",
                "reason": "网格计划",
            },
        ]
    )


def _readability_backtest_result() -> web_app.RenderResult:
    return web_app.RenderResult(
        "恒温器回测诊断",
        summaries=[
            {
                "backtest_type": "event_driven",
                "initial_cash": 100000.0,
                "final_value": 100234.567,
                "total_return": 0.012345,
            }
        ],
        tables=[
            web_app.TableBlock(
                "Trades",
                pd.DataFrame(
                    [
                        {
                            "symbol": "600519.SH",
                            "name": "贵州茅台",
                            "date": "2026-07-01",
                            "signal_time": "09:30",
                            "order_status": "filled",
                            "slippage_cost": 1.234,
                            "shares_after": 100,
                            "position_after": 100,
                            "price": 123.456,
                            "net_amount": 12345.678,
                        }
                    ]
                ),
            ),
            web_app.TableBlock(
                "Daily Portfolio",
                pd.DataFrame(
                    [
                        {
                            "date": "2026-07-01",
                            "total_value": 100234.567,
                            "cash": 87654.321,
                            "shares": 100,
                            "daily_return": 0.012345,
                        }
                    ]
                    * 55
                ),
            ),
            web_app.TableBlock(
                "Data Quality",
                pd.DataFrame(),
            ),
        ],
        extra_html='<a href="/reports/demo.xlsx">下载 Excel 报告</a>',
    )


def test_web_app_parses_symbols_and_marks() -> None:
    form = {"symbols": "600519, 000001\n600036", "marks": "600519=1500.5,000001=12.3"}

    assert web_app._symbols(form) == ["600519", "000001", "600036"]
    assert web_app._marks(form) == {"600519": 1500.5, "000001": 12.3}


def test_web_default_path_is_thermostat_and_hides_old_entries() -> None:
    html = web_app.render_page(page="unknown")

    assert "恒温器策略" in html
    assert 'action="/thermostat-job"' in html
    assert 'href="/thermostat"' in html
    assert 'href="/portfolio"' in html
    assert 'action="/turtle"' not in html
    assert 'action="/turtle-backtest"' not in html
    assert 'href="/turtle"' not in html
    assert "海龟系统" not in html
    assert "默认技术筛选" not in html
    assert 'href="/strategy"' not in html


def test_web_pages_use_workbench_shell() -> None:
    for page, title in [
        ("thermostat", "恒温器策略"),
        ("backtest", "恒温器回测诊断"),
        ("portfolio", "账户"),
    ]:
        html = web_app.render_page(page=page)

        assert 'class="workbench-page"' in html
        assert title in html
        assert "页面状态" in html
        assert "工作区" in html
        assert "海龟系统" not in html


def test_web_thermostat_page_shows_stock_pool_controls(tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("高关注")
    store.add_symbols("高关注", ["600519"])
    store.save_last_manual_input("600519,000001")

    html = web_app.render_page(
        page="thermostat",
        form={"account_path": str(tmp_path / "account"), "stock_pool_source": "watchlist"},
    )

    assert "股票池来源" in html
    assert 'name="stock_pool_source"' in html
    assert "手动输入" in html
    assert "自选股组合" in html
    assert "市场范围" in html
    assert "龙虎榜" in html
    assert "同花顺龙虎榜" not in html
    assert "高关注" in html
    assert "600519,000001" not in html
    assert "剔除科创板" in html
    assert "将在后续阶段接入" not in html
    assert "海龟系统" not in html


def test_web_thermostat_get_query_rerenders_selected_lhb_fields() -> None:
    html = web_app.render_page(page="thermostat", form={"stock_pool_source": "lhb", "lhb_range": "1w"})

    assert 'action="/thermostat-job"' in html
    assert "龙虎榜时间范围" in html
    assert "运行候选数量" in html
    assert "前 20 名" in html
    assert "前 30 名" in html
    assert "前 50 名" in html
    assert "编辑手动股票池" not in html


def test_web_thermostat_all_sources_submit_to_same_job_entry(tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])

    forms = [
        {"stock_pool_source": "manual", "symbols": "600519"},
        {"stock_pool_source": "watchlist", "account_path": str(tmp_path / "account"), "watchlist_name": "观察"},
        {"stock_pool_source": "market_range", "market_range": "all_a"},
        {"stock_pool_source": "lhb", "lhb_range": "1w", "lhb_confirmed_top": "30"},
    ]

    for form in forms:
        html = web_app.render_page(page="thermostat", form=form)
        assert '<form method="post" action="/thermostat-job">' in html
        assert '<form method="post" action="/thermostat">' not in html
        assert "运行恒温器策略" in html


def test_web_stock_pool_source_selector_refreshes_without_running() -> None:
    html = web_app.render_page(page="thermostat", form={"stock_pool_source": "manual"})

    assert 'name="stock_pool_source"' in html
    assert 'data-source-selector="stock_pool_source"' in html
    assert "refreshSourceFields" in html


def test_web_thermostat_uses_conditional_stock_pool_controls(tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("高关注")
    store.add_symbols("高关注", ["600519"])

    manual_html = web_app.render_page(
        page="thermostat",
        form={"account_path": str(tmp_path / "account"), "stock_pool_source": "manual", "symbols": "600519 000001"},
    )
    watchlist_html = web_app.render_page(
        page="thermostat",
        form={"account_path": str(tmp_path / "account"), "stock_pool_source": "watchlist"},
    )
    empty_watchlist_html = web_app.render_page(
        page="thermostat",
        form={"account_path": str(tmp_path / "empty"), "stock_pool_source": "watchlist"},
    )
    market_html = web_app.render_page(page="thermostat", form={"stock_pool_source": "market_range"})
    lhb_html = web_app.render_page(page="thermostat", form={"stock_pool_source": "lhb", "lhb_range": "1w"})
    custom_lhb_html = web_app.render_page(page="thermostat", form={"stock_pool_source": "lhb", "lhb_range": "custom"})

    assert "编辑手动股票池" in manual_html
    assert "仅本次使用" in manual_html
    assert "已识别股票数量" in manual_html
    assert "保存手动股票池" not in manual_html
    assert "自选股组合名称" not in manual_html
    assert "高关注" in watchlist_html
    assert "暂无自选组合，请到账户页创建" in empty_watchlist_html
    assert "沪深 A 股" in market_html
    assert "创业板" in market_html
    assert "北交所" in market_html
    assert 'type="checkbox" name="market_range"' in market_html
    assert "龙虎榜开始日期" not in lhb_html
    assert "龙虎榜结束日期" not in lhb_html
    assert 'action="/thermostat-job"' in lhb_html
    assert 'action="/thermostat-lhb-preview"' not in lhb_html
    assert 'name="lhb_confirmed_top"' in lhb_html
    assert "前 20 名" in lhb_html
    assert "前 30 名" in lhb_html
    assert "前 50 名" in lhb_html
    assert "龙虎榜开始日期" in custom_lhb_html
    assert "龙虎榜结束日期" in custom_lhb_html


def test_web_thermostat_uses_date_range_account_cash_and_advanced_settings(tmp_path) -> None:
    account_path = tmp_path / "account"
    ManualPortfolioStore(account_path).initialize(principal=100000, cash=87654)

    initialized_html = web_app.render_page(page="thermostat", form={"account_path": str(account_path)})
    missing_html = web_app.render_page(page="thermostat", form={"account_path": str(tmp_path / "missing")})
    simulated_html = web_app.render_page(page="thermostat", form={"use_simulated_cash": "on"})
    custom_range_html = web_app.render_page(page="thermostat", form={"strategy_date_range": "custom"})

    assert "策略日期范围" in initialized_html
    assert "最近 1 个月" in initialized_html
    assert "最近 3 个月" in initialized_html
    assert "最近半年" in initialized_html
    assert "最近 1 年" in initialized_html
    assert "策略开始日期" not in initialized_html
    assert "策略结束日期" not in initialized_html
    assert "策略开始日期" in custom_range_html
    assert "策略结束日期" in custom_range_html
    assert "账户现金" in initialized_html
    assert "87654" in initialized_html
    assert 'name="cash"' not in initialized_html
    assert "账户未初始化，请先到账户页初始化账户" in missing_html
    assert "使用模拟资金" in initialized_html
    assert 'name="cash"' not in initialized_html
    assert "临时策略测算" in simulated_html
    assert 'name="cash"' in simulated_html
    assert "高级设置" in initialized_html
    assert "数据与执行设置" in initialized_html


def test_thermostat_strategy_custom_date_range_refreshes_without_actual_range() -> None:
    default_html = web_app.render_page(page="thermostat", form={"strategy_date_range": "3m"})
    custom_html = web_app.render_page(
        page="thermostat",
        form={"strategy_date_range": "custom", "start": "20240101", "end": "20260706"},
    )

    assert 'name="strategy_date_range"' in default_html
    assert '<select name="strategy_date_range" onchange="refreshSourceFields(this)">' in default_html
    assert 'window.location.href = "/thermostat?"' in default_html
    assert "实际使用日期范围" in default_html
    assert "实际使用日期范围" not in custom_html
    assert 'name="start"' in custom_html
    assert 'name="end"' in custom_html
    assert 'value="20240101"' in custom_html
    assert 'value="20260706"' in custom_html


def test_web_can_save_manual_input_as_watchlist(tmp_path) -> None:
    result = web_app.handle_watchlist_save_manual(
        {
            "path": str(tmp_path / "account"),
            "symbols": "600519,000001",
            "watchlist_name": "高关注",
        }
    )

    saved = WatchlistStore(tmp_path / "account").get("高关注")

    assert saved is not None
    assert saved.symbols == ["600519.SH", "000001.SZ"]
    assert result.title == "自选股已保存"


def test_web_portfolio_page_uses_overview_and_function_tabs(tmp_path) -> None:
    account_path = tmp_path / "account"
    store = ManualPortfolioStore(account_path)
    store.initialize(principal=100000, cash=90000)
    store.buy("600001", name="A", price=10.0, shares=100, fees=0.0)
    for idx in range(6):
        store.adjust_cost("600001", avg_cost=10.0 + idx * 0.1, note=f"note-{idx}")

    html = web_app.render_page(page="portfolio", form={"path": str(account_path)})

    assert "账户概览" in html
    assert 'class="overview-card"' in html
    for label in ["本金", "现金", "持仓市值", "总资产", "总收益", "总收益率", "已实现盈亏", "浮动盈亏", "持仓数量", "胜率", "盈亏比", "最大回撤", "佣金率", "印花税率"]:
        assert label in html
    assert "当前持仓" in html
    assert "查看全部持仓" in html
    assert "交易流水" in html
    assert "查看全部交易流水" in html
    assert html.count("adjust_cost") <= 5
    assert "功能操作区" in html
    for tab in ["自选组合", "账户设置", "持仓与估值", "买入 / 卖出", "成本调整", "交易记录"]:
        assert tab in html
    assert "高级信息" in html
    assert "会修改持仓成本记录" in html


def test_web_portfolio_empty_states(tmp_path) -> None:
    uninitialized = web_app.render_page(page="portfolio", form={"path": str(tmp_path / "missing")})
    initialized_path = tmp_path / "account"
    ManualPortfolioStore(initialized_path).initialize(principal=100000)
    initialized = web_app.render_page(page="portfolio", form={"path": str(initialized_path)})

    assert "账户未初始化" in uninitialized
    assert "初始化账户" in uninitialized
    assert "暂无持仓" in initialized
    assert "暂无交易流水" in initialized
    assert "暂无自选组合" in initialized


def test_web_portfolio_page_exposes_watchlist_management_without_hiding_account_forms(tmp_path) -> None:
    WatchlistStore(tmp_path / "account").create("高关注")

    html = web_app.render_page(page="portfolio", form={"path": str(tmp_path / "account")})

    assert "自选股组合" in html
    assert "自选股不是持仓" in html
    assert 'action="/watchlist-create"' in html
    assert 'action="/watchlist-add-symbol"' in html
    assert 'action="/watchlist-remove-symbol"' in html
    assert 'action="/watchlist-rename"' in html
    assert 'action="/watchlist-delete"' in html
    assert "高关注" in html
    assert 'action="/portfolio-buy"' in html
    assert 'action="/portfolio-sell"' in html


def test_web_watchlist_add_symbol_splits_batch_input_and_reports_invalid(tmp_path) -> None:
    account_path = tmp_path / "account"
    store = WatchlistStore(account_path)
    store.create("短线观察")

    result = web_app.handle_watchlist_action(
        "/watchlist-add-symbol",
        {"path": str(account_path), "watchlist_name": "短线观察", "symbol": "600519, 000001 300750\n516650"},
    )
    saved = WatchlistStore(account_path).get("短线观察")
    html = web_app.render_message(result, None)

    assert saved is not None
    assert saved.symbols == ["600519.SH", "000001.SZ", "300750.SZ"]
    assert "516650" in html
    assert "无法识别或暂不支持" in html


def test_web_watchlist_table_flags_historical_invalid_symbols(tmp_path) -> None:
    account_path = tmp_path / "account"
    account_path.mkdir()
    (account_path / "watchlists.json").write_text(
        json.dumps(
            {
                "持仓": {
                    "symbols": ["516650,515880", "515070", "600519.SH"],
                    "updated_at": "2026-06-30T00:00:00",
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    html = web_app.render_page(page="portfolio", form={"path": str(account_path)})

    assert "存在异常代码" in html
    assert "516650" in html
    assert "515070" in html
    assert "未翻译字段" not in html
    assert "更新时间" in html


def test_web_normal_pages_do_not_show_old_strategy_entries() -> None:
    for page in ["thermostat", "backtest", "portfolio"]:
        html = web_app.render_page(page=page)
        _assert_no_mojibake(html, context=page)
        assert 'action="/turtle"' not in html
        assert 'action="/turtle-backtest"' not in html
        assert 'href="/turtle"' not in html
        assert "海龟系统" not in html
        assert "旧策略列表" not in html


def test_web_source_no_longer_keeps_unreachable_turtle_pages() -> None:
    source = web_app.Path(web_app.__file__).read_text(encoding="utf-8")

    assert "def render_turtle_section" not in source
    assert "def turtle_fields" not in source
    assert "def handle_turtle(" not in source
    assert "def handle_turtle_backtest" not in source
    assert "run_turtle_system" not in source
    assert "backtest_turtle_system" not in source
    assert "def _turtle_config" not in source
    assert "def _resolve_turtle_universe" not in source
    assert "def _resolve_backtest_universe" not in source


def test_web_job_progress_messages_are_readable() -> None:
    job = web_app.ThermostatJob("job-1", {"symbols": "600001"})

    queued = web_app.job_status_payload("job-1")
    assert queued["status"] == "missing"
    _assert_no_mojibake(str(queued), context="missing job")

    html = web_app.render_job_progress("job-1")
    _assert_no_mojibake(html, context="job progress html")

    job.update({"stage": "load_candidate_history", "completed": 3, "total": 5, "current_symbol": "600001.SH"})
    payload = {
        "node": job.node,
        "message": job.message,
        "stage": job.stage,
    }
    _assert_no_mojibake(str(payload), context="running job payload")
    assert "正在加载候选股历史" in job.node
    assert "已完成 3 / 5" in job.message


def test_web_thermostat_result_uses_trigger_plan_sections_without_legacy_advice(tmp_path, monkeypatch) -> None:
    store = ManualPortfolioStore(tmp_path / "account")
    store.initialize(principal=100000, cash=80000)
    store.buy("600001", name="Held", price=18.0, shares=100, fees=0.0, strategy="thermostat", system="risk_control")
    fake = FakeWebService(
        {
            "600001.SH": _history("600001.SH", [20 - i * 0.25 for i in range(40)]),
            "600002.SH": _history("600002.SH", [10 + i * 0.25 for i in range(40)]),
        },
        quotes=pd.DataFrame(
            [
                {"symbol": "600002.SH", "name": "Candidate", "price": 19.8, "high": 20.0, "prev_close": 19.0, "volume": 100000}
            ]
        ),
    )
    monkeypatch.setattr(web_app, "_service", lambda form: fake)

    result = web_app.handle_thermostat(
        {
            "symbols": "600002",
            "account_path": str(tmp_path / "account"),
            "start": "20260401",
            "end": "20260510",
            "execution_plan": "on",
        }
    )

    titles = [table.title for table in result.tables]
    assert titles == [
        "市场状态与仓位折扣",
        "个股模式摘要",
        "趋势触发计划",
        "网格触发计划",
        "待卖记录",
        "失败原因和风险提示",
        "错误/数据质量",
    ]
    for old_title in ["Holding Advice", "New Buy Candidates", "Grid Advice", "Trend Advice", "Execution Plan", "持仓建议", "新买候选", "网格建议", "趋势建议"]:
        assert old_title not in titles
    summary = next(table.frame for table in result.tables if table.title == "个股模式摘要")
    assert {"600001.SH", "600002.SH"}.issubset(set(summary["symbol"]))


def test_web_thermostat_result_renders_t1_trigger_plan_sections(tmp_path, monkeypatch) -> None:
    account_path = tmp_path / "account"
    ManualPortfolioStore(account_path).initialize(principal=100000, cash=90000)
    fake = FakeWebService({})
    monkeypatch.setattr(web_app, "_service", lambda form: fake)

    trigger_plan = pd.DataFrame(
        [
            {
                "symbol": "600001.SH",
                "name": "Trend",
                "date": "2025-05-20",
                "market_regime": "market_uptrend",
                "market_regime_normalized": "strong",
                "market_position_discount": 1.0,
                "stock_mode": "trend",
                "reference_price": 12.0,
                "boll_upper": 12.8,
                "boll_mid": 11.5,
                "boll_lower": 10.2,
                "atr20": 0.42,
                "trend_buy_trigger": 12.9,
                "trend_reduce_trigger": 11.5,
                "trend_exit_trigger": 10.0,
                "trend_batches": "40%,35%,25%",
                "target_position_pct": 0.2,
                "max_position_pct": 0.2,
                "total_shares": 100,
                "available_shares": 0,
                "today_bought_shares": 100,
                "pending_sell_level": "pending_exit",
                "trigger_status": "triggered",
                "filled_status": "pending",
                "failed_reason": "",
                "risk_note": "今日买入不可卖",
                "reason": "趋势计划",
            },
            {
                "symbol": "600002.SH",
                "name": "Range",
                "date": "2025-05-20",
                "market_regime": "market_range",
                "market_regime_normalized": "normal",
                "market_position_discount": 0.9,
                "stock_mode": "range",
                "reference_price": 9.8,
                "grid_lower": 9.1,
                "grid_mid": 10.0,
                "grid_upper": 10.9,
                "grid_max_layers": 3,
                "grid_buy_levels": "9.65|9.30|9.10",
                "grid_sell_levels": "10.35|10.70|10.90",
                "grid_total_max_position_pct": 0.4,
                "target_position_pct": 0.135,
                "max_position_pct": 0.135,
                "total_shares": 0,
                "available_shares": 0,
                "today_bought_shares": 0,
                "pending_sell_level": "",
                "trigger_status": "planned",
                "filled_status": "failed",
                "failed_reason": "limit_up_buy_failed",
                "risk_note": "",
                "reason": "网格计划",
            },
        ]
    )

    def fake_run_thermostat_strategy(**kwargs):
        return type(
            "FakeThermostatResult",
            (),
            {
                "market_overview": pd.DataFrame([{"market_regime": "market_uptrend"}]),
                "holding_advice": pd.DataFrame(),
                "new_candidates": pd.DataFrame(),
                "grid_advice": pd.DataFrame(),
                "trend_advice": pd.DataFrame(),
                "errors": pd.DataFrame(),
                "trigger_plan": trigger_plan,
            },
        )()

    monkeypatch.setattr(web_app, "run_thermostat_strategy", fake_run_thermostat_strategy)

    result = web_app.handle_thermostat(
        {
            "symbols": "600001,600002",
            "start": "20250101",
            "end": "20250520",
            "account_path": str(account_path),
        }
    )
    html = web_app.render_message(result, None)
    titles = [table.title for table in result.tables]

    for title in ["市场状态与仓位折扣", "个股模式摘要", "趋势触发计划", "网格触发计划", "待卖记录", "失败原因和风险提示", "错误/数据质量"]:
        assert title in titles
        assert title in html
    for old_title in ["Holding Advice", "New Buy Candidates", "Grid Advice", "Trend Advice", "Execution Plan", "持仓建议", "新买候选", "网格建议", "趋势建议", "手工执行计划"]:
        assert old_title not in titles
        assert old_title not in html
    assert "展开查看详细字段" in html
    assert "<details" in html
    assert "stock_mode" in html
    assert "trend_buy_trigger" in html
    assert "股票模式" in html
    assert "趋势买入触发价" in html
    assert "今日买入股数" in html

    trend_table = next(table.frame for table in result.tables if table.title == "趋势触发计划")
    grid_table = next(table.frame for table in result.tables if table.title == "网格触发计划")
    pending_table = next(table.frame for table in result.tables if table.title == "待卖记录")
    risk_table = next(table.frame for table in result.tables if table.title == "失败原因和风险提示")
    assert trend_table["symbol"].tolist() == ["600001.SH"]
    assert grid_table["symbol"].tolist() == ["600002.SH"]
    assert pending_table["symbol"].tolist() == ["600001.SH"]
    assert set(risk_table["symbol"]) == {"600001.SH", "600002.SH"}
    assert html.count("<th>stock_mode</th>") == 0


def test_web_thermostat_form_does_not_add_duplicate_account_inputs() -> None:
    html = web_app.render_page(page="thermostat", form={"account_path": "data/user/default"})

    assert html.count('name="account_path"') == 1
    assert 'name="available_shares"' not in html
    assert 'name="today_bought_shares"' not in html
    assert 'name="total_shares"' not in html
    assert 'name="commission_rate"' not in html
    assert 'name="stamp_tax_rate"' not in html


def test_web_thermostat_job_result_shows_t1_report_download_only_after_completion(tmp_path) -> None:
    initial_html = web_app.render_page(page="thermostat", form={})
    assert "下载新版 T+1 恒温器报告" not in initial_html

    job = web_app.ThermostatJob("report-job", {"symbols": "600001"})
    job.report_path = str(tmp_path / "t1_thermostat_report_20260708.xlsx")
    job.report_filename = "t1_thermostat_report_20260708.xlsx"
    result = web_app.RenderResult("恒温器策略", tables=[], summaries=[])

    job.complete(result)

    assert "下载新版 T+1 恒温器报告" in job.result_html
    assert "/thermostat-report?id=report-job" in job.result_html
    for old_title in ["Holding Advice", "New Buy Candidates", "Grid Advice", "Trend Advice", "持仓建议", "新买候选", "网格建议", "趋势建议"]:
        assert old_title not in job.result_html


def test_web_thermostat_job_result_shows_report_export_failure() -> None:
    job = web_app.ThermostatJob("failed-report-job", {"symbols": "600001"})
    job.report_error = "disk full"

    job.complete(web_app.RenderResult("恒温器策略", tables=[], summaries=[]))

    assert "报告导出失败" in job.result_html
    assert "disk full" in job.result_html
    assert "下载新版 T+1 恒温器报告" not in job.result_html


def test_web_thermostat_report_route_downloads_current_t1_excel(tmp_path) -> None:
    report = build_t1_thermostat_report(_sample_trigger_plan(), pd.DataFrame())
    output = tmp_path / "t1_thermostat_report_20260708.xlsx"
    export_t1_thermostat_excel(report, output)
    job = web_app.ThermostatJob("download-job", {"symbols": "600001,600002"})
    job.status = "done"
    job.report_path = str(output)
    job.report_filename = output.name
    web_app.JOBS["download-job"] = job
    server = web_app.ThreadingHTTPServer(("127.0.0.1", 0), web_app.WebAppHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        port = server.server_address[1]
        request = Request(f"http://127.0.0.1:{port}/thermostat-report?id=download-job")
        with urlopen(request, timeout=10) as response:
            body = response.read()
            content_type = response.headers["Content-Type"]
            disposition = response.headers["Content-Disposition"]
    finally:
        server.shutdown()
        thread.join(timeout=5)
        web_app.JOBS.pop("download-job", None)

    downloaded = tmp_path / "downloaded.xlsx"
    downloaded.write_bytes(body)
    book = load_workbook(downloaded)
    assert response.status == 200
    assert content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="t1_thermostat_report_20260708.xlsx"' in disposition
    assert "个股模式摘要" in book.sheetnames
    assert "详细字段" in book.sheetnames
    assert not {"Holding Advice", "New Buy Candidates", "Grid Advice", "Trend Advice"} & set(book.sheetnames)


def test_web_thermostat_job_runner_generates_t1_excel_report(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(web_app, "REPORT_DIR", tmp_path)
    result = web_app.RenderResult(
        "恒温器策略",
        tables=[],
        summaries=[],
        metadata={"trigger_plan": _sample_trigger_plan(), "errors": pd.DataFrame()},
    )
    monkeypatch.setattr(web_app, "handle_thermostat", lambda form, progress_callback=None: result)
    job = web_app.ThermostatJob("runner-report-job", {"symbols": "600001,600002"})
    web_app.JOBS["runner-report-job"] = job

    try:
        web_app._run_thermostat_job("runner-report-job")
    finally:
        web_app.JOBS.pop("runner-report-job", None)

    assert job.status == "done"
    assert job.report_filename == "t1_thermostat_report_20260708.xlsx"
    assert Path(job.report_path).exists()
    assert "下载新版 T+1 恒温器报告" in job.result_html
    book = load_workbook(job.report_path)
    assert "详细字段" in book.sheetnames


def test_web_backtest_job_builds_21_sheet_report_from_identical_raw_result(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(web_app, "REPORT_DIR", tmp_path)
    raw = _minimal_backtest_result()
    raw.daily_assets = pd.DataFrame(
        [{"date": "2026-07-13", "cash": 100000.0, "total_asset": 100000.0}]
    )
    before = {name: frame.copy(deep=True) for name, frame in raw.tables.items()}
    rendered = web_app.RenderResult(
        "T1 恒温器回测",
        metadata={"backtest_result": raw},
    )
    monkeypatch.setattr(
        web_app,
        "handle_thermostat_backtest",
        lambda form, progress_callback=None: rendered,
    )
    job = web_app.ThermostatJob("backtest-report-job", {"symbols": "600001"})
    web_app.JOBS[job.job_id] = job

    try:
        web_app._run_thermostat_backtest_job(job.job_id)
    finally:
        web_app.JOBS.pop(job.job_id, None)

    assert job.status == "done"
    assert job.report_type == "t1_backtest"
    assert re.fullmatch(
        r"t1_thermostat_backtest_\d{8}_\d{6}\.xlsx",
        job.report_filename,
    )
    assert Path(job.report_path).exists()
    assert "/thermostat-backtest-report?id=backtest-report-job" in job.result_html
    assert "下载 T+1 恒温器回测报告" in job.result_html
    assert load_workbook(job.report_path).sheetnames == EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS
    for name, frame in raw.tables.items():
        pd.testing.assert_frame_equal(frame, before[name])


def test_web_backtest_report_failure_is_visible_and_raw_result_remains_renderable(monkeypatch) -> None:
    raw = _minimal_backtest_result()
    rendered = web_app.RenderResult(
        "T1 恒温器回测",
        metadata={"backtest_result": raw},
    )
    monkeypatch.setattr(
        web_app,
        "handle_thermostat_backtest",
        lambda form, progress_callback=None: rendered,
    )
    monkeypatch.setattr(
        web_app,
        "build_t1_thermostat_backtest_report",
        lambda result: (_ for _ in ()).throw(OSError("disk full")),
    )
    job = web_app.ThermostatJob("backtest-report-failed", {})
    web_app.JOBS[job.job_id] = job

    try:
        web_app._run_thermostat_backtest_job(job.job_id)
    finally:
        web_app.JOBS.pop(job.job_id, None)

    assert job.status == "done"
    assert job.report_error == "disk full"
    assert "报告导出失败" in job.result_html
    assert "回测状态" in job.result_html
    assert rendered.metadata["backtest_result"] is raw


def test_completed_backtest_has_exactly_one_real_report_section_and_preparing_state(tmp_path) -> None:
    raw = _minimal_backtest_result()
    result = web_app.RenderResult(
        "T1 恒温器回测",
        metadata={"backtest_result": raw},
    )
    job = web_app.ThermostatJob("single-report-section", {})
    job.report_type = "t1_backtest"
    job.report_path = str(tmp_path / "t1_thermostat_backtest_20260713_090807.xlsx")
    job.report_filename = Path(job.report_path).name

    job.complete(result)

    assert job.result_html.count('class="result-section result-section-report"') == 1
    assert job.result_html.count("下载 T+1 恒温器回测报告") == 1
    assert "Excel 报告将在下一阶段提供" not in job.result_html
    assert "disabled" not in job.result_html

    preparing = web_app.ThermostatJob("preparing-report", {})
    preparing.report_type = "t1_backtest"
    preparing_html = web_app._thermostat_report_entry(preparing)
    assert preparing_html.count('class="result-section result-section-report"') == 1
    assert "正在准备" in preparing_html
    assert "disabled" not in preparing_html


def test_web_backtest_report_route_is_type_safe_and_keeps_strategy_route(tmp_path) -> None:
    backtest_path = tmp_path / "t1_thermostat_backtest_20260713_090807.xlsx"
    backtest_path.write_bytes(b"backtest-xlsx")
    strategy_path = tmp_path / "t1_thermostat_report_20260713.xlsx"
    strategy_path.write_bytes(b"strategy-xlsx")
    backtest_job = web_app.ThermostatJob("typed-backtest", {})
    backtest_job.status = "done"
    backtest_job.report_type = "t1_backtest"
    backtest_job.report_path = str(backtest_path)
    backtest_job.report_filename = backtest_path.name
    strategy_job = web_app.ThermostatJob("typed-strategy", {})
    strategy_job.status = "done"
    strategy_job.report_type = "t1_strategy"
    strategy_job.report_path = str(strategy_path)
    strategy_job.report_filename = strategy_path.name
    web_app.JOBS.update({backtest_job.job_id: backtest_job, strategy_job.job_id: strategy_job})
    server = web_app.ThreadingHTTPServer(("127.0.0.1", 0), web_app.WebAppHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        port = server.server_address[1]
        with urlopen(
            f"http://127.0.0.1:{port}/thermostat-backtest-report?id=typed-backtest",
            timeout=10,
        ) as response:
            assert response.read() == b"backtest-xlsx"
            assert response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        with urlopen(
            f"http://127.0.0.1:{port}/thermostat-report?id=typed-strategy",
            timeout=10,
        ) as response:
            assert response.read() == b"strategy-xlsx"
    finally:
        server.shutdown()
        thread.join(timeout=5)
        web_app.JOBS.pop(backtest_job.job_id, None)
        web_app.JOBS.pop(strategy_job.job_id, None)


def test_web_lhb_preview_builds_candidates_before_running_thermostat(monkeypatch) -> None:
    ranked = pd.DataFrame(
        [
            {"code": f"600{i:03d}", "name": f"Stock{i}", "net_buy": 1000 - i, "rank": i}
            for i in range(1, 61)
        ]
    )
    monkeypatch.setattr(web_app, "build_lhb_candidates", lambda start, end, top: (ranked.head(top), ranked))

    result = web_app.handle_thermostat_lhb_preview(
        {
            "stock_pool_source": "lhb",
            "lhb_range": "1w",
            "end": "20260629",
            "exclude_star": "on",
        }
    )

    assert result.title == "LHB Candidate Preview"
    summary = result.summaries[0]
    assert summary["actual_lhb_range"] == "20260623 至 20260629"
    assert summary["candidate_count"] == 60
    assert summary["top_options"] == "20 / 30 / 50"
    assert [table.title for table in result.tables] == ["LHB Top 20", "LHB Top 30", "LHB Top 50"]
    assert len(result.tables[0].frame) == 20
    assert len(result.tables[1].frame) == 30
    assert len(result.tables[2].frame) == 50


def test_web_lhb_source_runs_directly_after_range_and_top_selection(monkeypatch) -> None:
    ranked = pd.DataFrame(
        [
            {"code": f"600{i:03d}", "name": f"A{i}", "net_buy": 1000 - i, "rank": i}
            for i in range(1, 61)
        ]
    )
    captured: dict[str, object] = {}
    monkeypatch.setattr(web_app, "build_lhb_candidates", lambda start, end, top: (ranked.head(top), ranked))

    def fake_run_thermostat_strategy(**kwargs):
        captured.update(kwargs)
        return type(
            "FakeThermostatResult",
            (),
            {
                "market_overview": pd.DataFrame([{"market_regime": "uptrend"}]),
                "holding_advice": pd.DataFrame(),
                "new_candidates": pd.DataFrame(),
                "grid_advice": pd.DataFrame(),
                "trend_advice": pd.DataFrame(),
                "errors": pd.DataFrame(),
            },
        )()

    monkeypatch.setattr(web_app, "run_thermostat_strategy", fake_run_thermostat_strategy)

    result = web_app.handle_thermostat(
        {
            "stock_pool_source": "lhb",
            "lhb_range": "1w",
            "lhb_confirmed_top": "20",
            "end": "20260629",
        }
    )

    assert result.title == "恒温器策略"
    assert len(captured["symbols"]) == 20
    assert captured["symbols"][:3] == ["600001.SH", "600002.SH", "600003.SH"]


def test_web_thermostat_progress_tracks_nodes_and_stock_counts() -> None:
    fake = FakeWebService(
        {
            "600001.SH": _history("600001.SH", [10 + i * 0.2 for i in range(40)]),
            "600002.SH": _history("600002.SH", [20 + i * 0.1 for i in range(40)]),
        }
    )
    events: list[dict[str, object]] = []

    web_app.run_thermostat_strategy(
        service=fake,
        symbols=["600001", "600002"],
        start_date="20260401",
        end_date="20260510",
        cash=100000,
        progress_callback=events.append,
    )

    assert {
        "stage": "load_candidate_history",
        "completed": 1,
        "total": 2,
        "current_symbol": "600001.SH",
        "node": "加载候选股历史",
    } in events
    assert {
        "stage": "evaluate_candidates",
        "completed": 2,
        "total": 2,
        "current_symbol": "600002.SH",
        "node": "评估候选股",
    } in events


def test_web_result_rendering_localizes_titles_columns_and_values() -> None:
    result = web_app.RenderResult(
        "恒温器策略",
        summaries=[
            {
                "stock_pool_source": "watchlist",
                "watchlist_name": "观察",
                "time_range": "1w",
                "data_sufficient": True,
            }
        ],
        tables=[
            web_app.TableBlock(
                "Stock Pool Summary",
                pd.DataFrame(
                    [
                        {
                            "stock_pool_source": "watchlist",
                            "watchlist_name": "观察",
                            "time_range": "1w",
                            "source_detail": "data/user/default",
                            "raw_count": 2,
                            "deduped_count": 2,
                            "filtered_count": 2,
                            "excluded_count": 0,
                        }
                    ]
                ),
            ),
            web_app.TableBlock(
                "Market Overview",
                pd.DataFrame(
                    [
                        {
                            "market_regime": "range",
                            "confidence": "medium",
                            "data_source": "index_history",
                            "data_sufficient": True,
                        }
                    ]
                ),
            ),
            web_app.TableBlock(
                "Grid Advice",
                pd.DataFrame([{"strategy_family": "grid", "action": "wait_confirm", "stock_regime": "downtrend"}]),
            ),
            web_app.TableBlock(
                "Trend Advice",
                pd.DataFrame([{"strategy_family": "trend_following", "action": "observe", "stock_regime": "uptrend"}]),
            ),
            web_app.TableBlock(
                "Execution Plan",
                pd.DataFrame(
                    [
                        {
                            "recommended_action": "buy",
                            "fallback_action": "switch_alternative",
                            "limit_status": "normal",
                            "volume_limit_pct": 0.05,
                            "skip_insufficient_cash": False,
                            "skip_volume_limit": True,
                        }
                    ]
                ),
            ),
        ],
    )

    html = web_app.render_message(result, None)

    for text in [
        "股票池摘要",
        "市场概览",
        "网格建议",
        "趋势建议",
        "手工执行计划",
        "自选组合名称",
        "时间范围",
        "来源说明",
        "原始数量",
        "被剔除数量",
        "市场状态",
        "置信度",
        "数据来源",
        "数据是否充足",
        "推荐操作",
        "备选操作",
        "涨跌停状态",
        "成交量限制比例",
        "资金不足跳过",
        "成交量限制跳过",
        "震荡区间",
        "上升趋势",
        "下降趋势",
        "趋势跟随",
        "网格策略",
        "观察",
        "买入",
        "等待确认",
        "是",
        "否",
    ]:
        assert text in html

    for forbidden in [
        "Stock Pool Summary",
        "Market Overview",
        "Grid Advice",
        "Trend Advice",
        "Execution Plan",
        "watchlist_name",
        "time_range",
        "source_detail",
        "raw_count",
        "excluded_count",
        "market_regime",
        "confidence",
        "data_source",
        "data_sufficient",
        "recommended_action",
        "fallback_action",
        "limit_status",
        "volume_limit_pct",
        "skip_insufficient_cash",
        "skip_volume_limit",
        ">range<",
        ">uptrend<",
        ">downtrend<",
        ">trend_following<",
        ">grid<",
        ">observe<",
        ">wait_confirm<",
    ]:
        assert forbidden not in html


def test_web_rendering_hides_unknown_user_visible_fields() -> None:
    html = web_app.render_table("Unknown Result", pd.DataFrame([{"unmapped_field": "abc"}]))

    assert "未翻译字段" not in html
    assert "unmapped_field" not in html
    assert "abc" not in html


def test_backtest_result_rendering_groups_sections_and_report_entry() -> None:
    html = web_app.render_message(_readability_backtest_result(), None)

    assert 'class="result-section result-section-summary"' in html
    assert 'class="result-section result-section-table"' in html
    assert 'class="report-entry report-entry-available"' in html
    assert "<summary>交易流水" in html
    assert "<summary>每日账户" in html
    assert "暂无数据。" in html
    assert "下载 Excel 报告" in html


def test_backtest_transaction_flow_hides_internal_fields_and_shows_stock_name() -> None:
    html = web_app.render_table("Trades", _readability_backtest_result().tables[0].frame)

    assert "股票" in html
    assert "名称" in html
    assert "贵州茅台" in html
    _assert_no_raw_internal_fields(html, ("signal_time", "order_status", "slippage_cost", "shares_after"))
    assert "filled" not in html
    assert "09:30" not in html


def test_backtest_result_money_values_use_two_decimals_without_mutating_other_types() -> None:
    html = web_app.render_message(_readability_backtest_result(), None)

    assert "100000.00" in html
    assert "100234.57" in html
    assert "123.46" in html
    assert "12345.68" in html
    assert "87,654.32" not in html
    assert "123.456000" not in html
    assert "2026-07-01" in html
    assert ">100<" in html
    assert "600519.SH" in html


def test_backtest_result_tables_keep_headers_visible_for_large_outputs() -> None:
    html = web_app.render_table("Daily Portfolio", _readability_backtest_result().tables[1].frame)

    assert 'class="table-wrap table-wrap-scroll"' in html
    assert 'class="data-table sticky-table"' in html
    assert "position: sticky" in web_app.CSS


def test_backtest_job_runner_emits_non_binary_progress_stages(monkeypatch) -> None:
    events: list[dict[str, object]] = []

    class FakeJob:
        form = {"symbols": "600001,600002", "start": "20260701", "end": "20260703"}

        def update(self, event: dict[str, object]) -> None:
            events.append(event)

        def complete(self, result: web_app.RenderResult, progress_callback=None) -> None:
            progress_callback({"stage": "prepare_report", "completed": 0, "total": 1})
            progress_callback({"stage": "prepare_report", "completed": 1, "total": 1})
            events.append({"stage": "done"})

        def fail(self, exc: Exception) -> None:
            raise exc

    monkeypatch.setitem(web_app.JOBS, "backtest-progress", FakeJob())
    def fake_handle(form, progress_callback=None):
        for stage in (
            "parse_backtest_request",
            "load_backtest_data",
            "simulate_daily",
            "calculate_metrics",
        ):
            progress_callback({"stage": stage, "completed": 2, "total": 2})
        return _readability_backtest_result()

    monkeypatch.setattr(web_app, "handle_thermostat_backtest", fake_handle)

    web_app._run_thermostat_backtest_job("backtest-progress")

    stages = [event.get("stage") for event in events]
    assert stages[:6] == [
        "parse_backtest_request",
        "load_backtest_data",
        "simulate_daily",
        "calculate_metrics",
        "prepare_report",
        "prepare_report",
    ]
    assert all(event.get("total", 0) != 1 for event in events[:4])
    assert events[4:6] == [
        {"stage": "prepare_report", "completed": 0, "total": 1},
        {"stage": "prepare_report", "completed": 1, "total": 1},
    ]


def test_backtest_result_display_formatting_does_not_mutate_raw_values() -> None:
    result = _readability_backtest_result()
    raw_summary_value = result.summaries[0]["final_value"]
    raw_trade_price = result.tables[0].frame.loc[0, "price"]
    raw_daily_value = result.tables[1].frame.loc[0, "total_value"]

    html = web_app.render_message(result, None)

    assert "100234.57" in html
    assert result.summaries[0]["final_value"] == raw_summary_value
    assert result.tables[0].frame.loc[0, "price"] == raw_trade_price
    assert result.tables[1].frame.loc[0, "total_value"] == raw_daily_value


def test_web_cash_shortfall_wording_distinguishes_suggested_position_from_account_cash() -> None:
    html = web_app.render_table(
        "New Buy Candidates",
        pd.DataFrame(
            [
                {
                    "symbol": "688135.SH",
                    "action": "observe",
                    "reason": "市场过渡期，仅试探仓；现金不足以买入一手",
                    "risk_note": "现金不足以买入一手",
                }
            ]
        ),
    )

    assert "建议仓位金额不足以买入一手" in html
    assert "账户现金不足" not in html


def test_web_job_progress_uses_chinese_stage_fallback_and_failure_summary() -> None:
    job = web_app.ThermostatJob("job-localization", {})

    job.update({"stage": "evaluate_candidates", "completed": 12, "total": 50, "current_symbol": "600519.SH"})

    assert job.node == "正在评估候选股"
    assert "已完成 12 / 50" in job.message
    assert "当前处理 600519.SH" in job.message
    assert "生成市场状态、网格/趋势建议" in job.message
    assert "evaluate_candidates" not in job.message

    job.fail(RuntimeError("upstream timeout"))

    assert job.error.startswith("任务失败：")
    assert "upstream timeout" in job.error


def test_web_normal_results_do_not_show_untranslated_field_marker(monkeypatch) -> None:
    fake = FakeWebService({"600001.SH": _history("600001.SH", [10 + i * 0.1 for i in range(80)])})
    monkeypatch.setattr(web_app, "_service", lambda form: fake)
    monkeypatch.setattr(
        web_app,
        "build_lhb_candidates",
        lambda start, end, top: (
            pd.DataFrame([{"code": "600001", "name": "A", "net_buy": 1000, "rank": 1}]),
            pd.DataFrame([{"code": "600001", "name": "A", "net_buy": 1000, "rank": 1}]),
        ),
    )
    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", lambda request, progress_callback=None: _minimal_backtest_result())

    results = [
        web_app.handle_thermostat(
            {
                "symbols": "600001",
                "start": "20260401",
                "end": "20260510",
                "account_path": "data/user/default",
            }
        ),
        web_app.handle_thermostat_lhb_preview({"stock_pool_source": "lhb", "lhb_range": "1w", "end": "20260629"}),
        web_app.handle_thermostat_job({"symbols": "600001"}),
        web_app.handle_thermostat_backtest({"symbols": "600001", "start": "20260401", "end": "20260510", "cash": "100000"}),
    ]

    for result in results:
        html = web_app.render_message(result, None)
        assert "未翻译字段" not in html
        assert ">queued<" not in html


def test_web_thermostat_rejects_invalid_and_empty_stock_pool(tmp_path) -> None:
    account_path = str(tmp_path / "missing-account")
    empty = web_app.handle_thermostat({"symbols": " ", "account_path": account_path})
    invalid = web_app.handle_thermostat({"symbols": "abc", "account_path": account_path})

    assert empty.title == "股票池错误"
    assert "手动输入为空" in empty.summaries[0]["errors"]
    assert invalid.title == "股票池错误"
    assert "abc" in invalid.summaries[0]["warnings"]


def test_web_thermostat_backtest_outputs_concise_t1_result(monkeypatch) -> None:
    fake = FakeWebService({"600001.SH": _history("600001.SH", [10 + i * 0.1 for i in range(80)])})
    monkeypatch.setattr(web_app, "_service", lambda form: fake)
    raw = _minimal_backtest_result()
    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", lambda request, progress_callback=None: raw)

    result = web_app.handle_thermostat_backtest(
        {"symbols": "600001", "start": "20260101", "end": "20260320", "cash": "100000"}
    )

    assert result.tables == []
    assert result.metadata["backtest_result"] is raw
    assert "核心指标" in web_app.render_message(result, None)


def test_backtest_page_shows_cache_parameters_results_and_download_sections() -> None:
    html = web_app.render_thermostat_backtest_section({})

    assert "数据缓存区" in html
    assert "回测参数区" in html
    assert "回测结果区" in html
    assert "报告下载区" in html


def test_backtest_page_uses_stock_pool_source_selector_for_manual_input() -> None:
    html = web_app.render_thermostat_backtest_section({"stock_pool_source": "manual", "symbols": "600519 000001"})

    assert 'name="stock_pool_source"' in html
    assert "编辑手动股票池" in html
    assert "手动股票池" in html
    assert "已识别股票数量" in html
    assert 'name="symbols"' in html


def test_backtest_page_uses_watchlist_dropdown_when_watchlists_exist(tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])

    html = web_app.render_thermostat_backtest_section(
        {"account_path": str(tmp_path / "account"), "stock_pool_source": "watchlist"}
    )

    assert "自选股组合" in html
    assert 'name="watchlist_name"' in html
    assert "观察" in html
    assert "手动股票池" not in html


def test_backtest_page_shows_empty_watchlist_state(tmp_path) -> None:
    html = web_app.render_thermostat_backtest_section(
        {"account_path": str(tmp_path / "empty"), "stock_pool_source": "watchlist"}
    )

    assert "暂无自选组合，请到账户页创建" in html
    assert "手动股票池" not in html


def test_backtest_page_uses_market_range_controls() -> None:
    html = web_app.render_thermostat_backtest_section({"stock_pool_source": "market_range"})

    assert "市场范围" in html
    assert "沪深 A 股" in html
    assert "创业板" in html
    assert "科创板" in html
    assert 'type="checkbox" name="market_range"' in html


def test_backtest_page_does_not_duplicate_lhb_candidate_sources() -> None:
    html = web_app.render_thermostat_backtest_section({"stock_pool_source": "lhb", "lhb_range": "1w"})

    assert "龙虎榜" in html
    assert "同花顺龙虎榜" not in html


def test_backtest_page_includes_date_range_presets() -> None:
    html = web_app.render_thermostat_backtest_section({"backtest_date_range": "5m", "end": "20260702"})

    assert 'name="backtest_date_range"' in html
    assert "最近 1 个月" in html
    assert "最近 3 个月" in html
    assert "最近 5 个月" in html
    assert "最近半年" in html
    assert "最近 1 年" in html
    assert "自定义" in html


def test_backtest_page_refreshes_conditional_controls_without_running() -> None:
    html = web_app.render_page(page="backtest", form={"stock_pool_source": "manual", "backtest_date_range": "3m"})

    assert "refreshSourceFields" in html
    assert 'data-source-selector="stock_pool_source"' in html
    assert 'name="backtest_date_range"' in html
    assert 'onchange="refreshSourceFields(this)"' in html
    assert 'window.location.href = "/backtest?"' in html


def test_backtest_form_submits_to_job_progress_endpoint() -> None:
    html = web_app.render_thermostat_backtest_section({"stock_pool_source": "watchlist", "backtest_date_range": "3m"})

    assert 'action="/thermostat-backtest-job"' in html
    assert 'action="/thermostat-backtest"' not in html


def test_backtest_job_start_renders_progress(monkeypatch) -> None:
    started: dict[str, object] = {}

    class FakeJob:
        job_id = "backtest-job-1"
        stage = "queued"
        node = "排队"
        message = "任务已创建，等待开始。"

    def fake_start(form):
        started.update(form)
        return FakeJob()

    monkeypatch.setattr(web_app, "start_thermostat_backtest_job", fake_start)

    result = web_app.handle_thermostat_backtest_job(
        {"stock_pool_source": "watchlist", "watchlist_name": "观察", "backtest_date_range": "3m"}
    )

    assert started["stock_pool_source"] == "watchlist"
    assert result.title == "恒温器回测任务已开始"
    assert 'data-job-id="backtest-job-1"' in result.extra_html
    assert "<progress" in result.extra_html


def test_backtest_page_preset_dates_show_summary_without_raw_date_inputs() -> None:
    html = web_app.render_thermostat_backtest_section(
        {"backtest_date_range": "5m", "start": "19990101", "end": "20260702"}
    )

    assert "实际回测日期范围" in html
    assert "20260202 至 20260702" in html
    assert 'name="start"' not in html
    assert 'name="end"' not in html


def test_backtest_page_custom_dates_show_editable_inputs() -> None:
    html = web_app.render_thermostat_backtest_section(
        {"backtest_date_range": "custom", "start": "20260101", "end": "20260702"}
    )

    assert "实际回测日期范围" not in html
    assert 'name="start"' in html
    assert 'name="end"' in html
    assert 'value="20260101"' in html
    assert 'value="20260702"' in html


def test_backtest_five_month_range_resolves_from_selected_end_date() -> None:
    assert web_app._backtest_range_dates({"backtest_date_range": "5m", "end": "20260702"}) == (
        "20260202",
        "20260702",
    )


def test_backtest_preset_ignores_stale_custom_dates() -> None:
    assert web_app._backtest_range_dates(
        {"backtest_date_range": "3m", "start": "19990101", "end": "20260702"}
    ) == ("20260403", "20260702")


def test_backtest_handler_resolves_watchlist_symbols_before_running(monkeypatch, tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])
    captured: dict[str, object] = {}

    def fake_backtest(request, progress_callback=None):
        captured["request"] = request
        return _minimal_backtest_result()

    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", fake_backtest)
    monkeypatch.setattr(web_app, "_service", lambda form: FakeWebService())

    result = web_app.handle_thermostat_backtest(
        {
            "account_path": str(tmp_path / "account"),
            "stock_pool_source": "watchlist",
            "watchlist_name": "观察",
            "symbols": "000001",
            "backtest_date_range": "5m",
            "end": "20260702",
            "cash": "100000",
        }
    )

    request = captured["request"]
    assert request.symbols == ("600519.SH",)
    assert request.start == "2026-02-02"
    assert request.end == "2026-07-02"
    assert result.metadata["backtest_result"] is not None


def test_backtest_handler_blocks_empty_watchlist_without_running(monkeypatch, tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("空组合")
    called = False

    def fake_backtest(request, progress_callback=None):
        nonlocal called
        called = True
        return _minimal_backtest_result()

    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", fake_backtest)

    result = web_app.handle_thermostat_backtest(
        {
            "account_path": str(tmp_path / "account"),
            "stock_pool_source": "watchlist",
            "watchlist_name": "空组合",
            "backtest_date_range": "3m",
            "end": "20260702",
        }
    )

    assert result.title == "股票池错误"
    assert called is False
    assert "空组合" in result.summaries[0]["errors"]


def test_backtest_handler_ignores_inactive_manual_symbols(monkeypatch, tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])
    captured: dict[str, object] = {}

    def fake_backtest(request, progress_callback=None):
        captured["request"] = request
        return _minimal_backtest_result()

    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", fake_backtest)
    monkeypatch.setattr(web_app, "_service", lambda form: FakeWebService())

    web_app.handle_thermostat_backtest(
        {
            "account_path": str(tmp_path / "account"),
            "stock_pool_source": "watchlist",
            "watchlist_name": "观察",
            "symbols": "000001",
            "backtest_date_range": "3m",
            "end": "20260702",
        }
    )

    assert captured["request"].symbols == ("600519.SH",)


def test_backtest_result_summary_uses_active_source_and_resolved_dates(monkeypatch, tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])

    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", lambda request, progress_callback=None: _minimal_backtest_result())
    monkeypatch.setattr(web_app, "_service", lambda form: FakeWebService())

    result = web_app.handle_thermostat_backtest(
        {
            "account_path": str(tmp_path / "account"),
            "stock_pool_source": "watchlist",
            "watchlist_name": "观察",
            "backtest_date_range": "5m",
            "end": "20260702",
        }
    )

    assert result.metadata["stock_pool_metadata"]["pool_type"] == "watchlist"
    request = result.metadata["data_request"]
    assert request.start == "2026-02-02"
    assert request.end == "2026-07-02"


def test_backtest_handler_rejects_unsupported_candidate_source_without_fallback(monkeypatch) -> None:
    called = False

    def fake_backtest(request, progress_callback=None):
        nonlocal called
        called = True
        return _minimal_backtest_result()

    monkeypatch.setattr(web_app, "run_t1_thermostat_backtest", fake_backtest)

    result = web_app.handle_thermostat_backtest(
        {
            "stock_pool_source": "ths_lhb",
            "symbols": "600519",
            "backtest_date_range": "3m",
            "end": "20260702",
        }
    )

    assert result.title == "股票池错误"
    assert called is False
    assert "不支持" in result.summaries[0]["errors"]


def test_thermostat_page_keeps_existing_stock_pool_workflow(tmp_path) -> None:
    store = WatchlistStore(tmp_path / "account")
    store.create("观察")
    store.add_symbols("观察", ["600519"])

    html = web_app.render_page(
        page="thermostat",
        form={"account_path": str(tmp_path / "account"), "stock_pool_source": "watchlist"},
    )

    assert 'action="/thermostat-job"' in html
    assert 'name="stock_pool_source"' in html
    assert 'data-source-selector="stock_pool_source"' in html
    assert 'name="watchlist_name"' in html
    assert "观察" in html
    assert 'action="/thermostat-backtest"' not in html


def test_account_page_keeps_watchlist_management_workflow(tmp_path) -> None:
    html = web_app.render_page(page="portfolio", form={"path": str(tmp_path / "account")})

    assert "自选组合" in html
    assert 'action="/watchlist-create"' in html
    assert 'action="/watchlist-add-symbol"' in html
    assert 'action="/watchlist-remove-symbol"' in html
    assert 'action="/watchlist-rename"' in html
    assert 'action="/watchlist-delete"' in html
    assert 'name="stock_pool_source"' not in html


def test_portfolio_trade_form_is_cleared_after_record() -> None:
    cleaned = web_app._clear_trade_form(
        {
            "path": "data/user/custom",
            "symbol": "600487",
            "price": "96.66",
            "shares": "100",
            "strategy_meta": "thermostat",
            "system": "trend_following",
            "realtime_source": "sina",
        }
    )

    html = web_app.render_page(page="portfolio", form=cleaned)

    assert cleaned == {"path": "data/user/custom", "realtime_source": "sina"}
    assert 'value="600487"' not in html
    assert 'value="96.66"' not in html
    assert 'value="100"' not in html
    assert 'value="thermostat"' in html
    assert 'value="trend_following"' in html


def test_web_display_form_after_success_clears_watchlist_inputs() -> None:
    base = {
        "path": "data/user/default",
        "account_path": "data/user/other",
        "watchlist_name": "观察",
        "symbol": "600519,000001",
        "new_watchlist_name": "新观察",
        "realtime_source": "sina",
    }

    created = web_app._display_form_after_success("/watchlist-create", base)
    added = web_app._display_form_after_success("/watchlist-add-symbol", base)
    removed = web_app._display_form_after_success("/watchlist-remove-symbol", base)
    renamed = web_app._display_form_after_success("/watchlist-rename", base)
    deleted = web_app._display_form_after_success("/watchlist-delete", base)

    assert created == {"path": "data/user/default", "account_path": "data/user/other"}
    assert added == {"path": "data/user/default", "account_path": "data/user/other", "watchlist_name": "观察"}
    assert removed == {"path": "data/user/default", "account_path": "data/user/other", "watchlist_name": "观察"}
    assert renamed == {"path": "data/user/default", "account_path": "data/user/other", "watchlist_name": "观察"}
    assert deleted == {"path": "data/user/default", "account_path": "data/user/other"}


def test_web_display_form_after_success_clears_portfolio_inputs() -> None:
    form = {
        "path": "data/user/default",
        "principal": "100000",
        "cash": "90000",
        "commission_rate": "0.0003",
        "min_commission": "5",
        "stamp_tax_rate": "0.001",
        "marks": "600519=1600",
        "source": "baostock",
        "stock_source": "akshare",
        "realtime_source": "sina",
        "refresh": "on",
    }

    initialized = web_app._display_form_after_success("/portfolio-init", form)
    refreshed = web_app._display_form_after_success("/portfolio-summary", {**form, "refresh_valuation": "1"})

    assert initialized == {"path": "data/user/default"}
    assert refreshed == {
        "path": "data/user/default",
        "source": "baostock",
        "stock_source": "akshare",
        "realtime_source": "sina",
        "refresh": "on",
    }


def test_web_page_form_state_is_isolated_by_page() -> None:
    mixed = {
        "path": "data/user/default",
        "account_path": "data/user/strategy",
        "symbol": "600519",
        "symbols": "000001",
        "watchlist_name": "观察",
        "stock_pool_source": "watchlist",
        "strategy_date_range": "3m",
        "backtest_date_range": "3m",
        "cash": "100000",
        "start": "20260101",
        "end": "20260201",
    }

    thermostat = web_app._display_form_for_page("thermostat", mixed)
    portfolio = web_app._display_form_for_page("portfolio", mixed)
    backtest = web_app._display_form_for_page("backtest", mixed)

    assert "symbol" not in thermostat
    assert thermostat["symbols"] == "000001"
    assert thermostat["watchlist_name"] == "观察"
    assert "symbols" not in portfolio
    assert portfolio["path"] == "data/user/default"
    assert "stock_pool_source" not in portfolio
    assert backtest["symbols"] == "000001"
    assert backtest["stock_pool_source"] == "watchlist"
    assert backtest["watchlist_name"] == "观察"
    assert backtest["backtest_date_range"] == "3m"


def test_web_account_path_normalization_prefers_account_path_for_thermostat() -> None:
    form = {"path": "data/user/account-page", "account_path": "data/user/strategy-page"}

    assert web_app._account_path_for_form(form, page="thermostat") == "data/user/strategy-page"
    assert web_app._account_path_for_form(form, page="portfolio") == "data/user/account-page"


def test_portfolio_summary_refreshes_valuation_without_adding_trades(tmp_path, monkeypatch) -> None:
    store = ManualPortfolioStore(tmp_path / "account")
    store.initialize(principal=100000)
    store.buy("600001", name="A", price=10.0, shares=100, fees=0.0)
    fake = FakeWebService(quotes=pd.DataFrame([{"symbol": "600001.SH", "name": "A", "price": 12.0}]))
    monkeypatch.setattr(web_app, "_service", lambda form: fake)

    result = web_app.handle_portfolio_summary({"path": str(tmp_path / "account"), "refresh_valuation": "1"})
    portfolio = store.load()

    assert result.summaries[0]["position_value"] == 1200.0
    assert len(portfolio.trades) == 1
    assert "mark_price" in result.tables[0].frame.columns


def test_web_app_adjusts_portfolio_cost(tmp_path) -> None:
    store = ManualPortfolioStore(tmp_path / "account")
    store.initialize(principal=100000)
    store.buy("002579", name="A", price=16.922, shares=100, fees=5.0)

    result = web_app.handle_portfolio_adjust_cost(
        {"path": str(tmp_path / "account"), "symbol": "002579", "avg_cost": "19.922"}
    )

    positions = result.tables[0].frame
    trades = result.tables[1].frame
    assert positions.loc[0, "avg_cost"] == 19.922
    assert trades.iloc[-1]["side"] == "adjust_cost"


def test_portfolio_table_html_contains_position_data(tmp_path) -> None:
    store = ManualPortfolioStore(tmp_path / "account")
    store.initialize(principal=100000)
    portfolio = store.buy("600001", name="A", price=10.0, shares=100, fees=0.0)

    html = web_app.render_table("Positions", web_app._positions_view(portfolio.positions))

    assert "600001.SH" in html
    assert "<tbody>" in html
