from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment


EXPECTED_T1_THERMOSTAT_SHEETS = [
    "市场状态与仓位折扣",
    "个股模式摘要",
    "趋势触发计划",
    "网格触发计划",
    "待卖记录",
    "失败原因和风险提示",
    "错误与数据质量",
    "详细字段",
]

TEXT_COLUMNS = {"grid_buy_levels", "grid_sell_levels", "failed_reason", "risk_note", "reason"}
PERCENT_COLUMNS = {
    "market_position_discount",
    "target_position_pct",
    "max_position_pct",
    "grid_total_max_position_pct",
}
INTEGER_COLUMNS = {"total_shares", "available_shares", "today_bought_shares", "grid_max_layers"}
DECIMAL_COLUMNS = {
    "reference_price",
    "boll_upper",
    "boll_mid",
    "boll_lower",
    "atr20",
    "trend_buy_trigger",
    "trend_reduce_trigger",
    "trend_exit_trigger",
    "grid_lower",
    "grid_mid",
    "grid_upper",
}

COLUMN_LABELS = {
    "date": "日期",
    "symbol": "股票",
    "code": "代码",
    "name": "名称",
    "stock_mode": "股票模式",
    "market_regime": "市场状态",
    "market_regime_normalized": "市场折扣档位",
    "market_position_discount": "仓位折扣",
    "target_position_pct": "目标仓位",
    "max_position_pct": "单股仓位上限",
    "total_shares": "总股数",
    "available_shares": "可卖股数",
    "today_bought_shares": "今日买入股数",
    "pending_sell_level": "待卖级别",
    "reference_price": "参考价",
    "boll_upper": "布林上轨",
    "boll_mid": "布林中轨",
    "boll_lower": "布林下轨",
    "atr20": "ATR20",
    "trend_buy_trigger": "趋势买入触发价",
    "trend_reduce_trigger": "趋势减仓触发价",
    "trend_exit_trigger": "趋势退出触发价",
    "trend_batches": "趋势分批",
    "grid_lower": "网格下沿",
    "grid_mid": "网格中枢",
    "grid_upper": "网格上沿",
    "grid_max_layers": "最大网格层数",
    "grid_buy_levels": "网格买入层",
    "grid_sell_levels": "网格卖出层",
    "grid_total_max_position_pct": "网格总仓位上限",
    "trigger_status": "触发状态",
    "filled_status": "成交状态",
    "failed_reason": "失败原因",
    "risk_note": "风险提示",
    "reason": "原因",
    "stock_regime": "原始股票状态",
    "share_split_source": "股数拆分来源",
    "data_sufficient": "数据是否充足",
    "error": "错误",
}

MARKET_COLUMNS = ["date", "market_regime", "market_regime_normalized", "market_position_discount"]
SUMMARY_COLUMNS = [
    "date",
    "symbol",
    "name",
    "stock_mode",
    "market_regime",
    "market_regime_normalized",
    "market_position_discount",
    "target_position_pct",
    "max_position_pct",
    "total_shares",
    "available_shares",
    "today_bought_shares",
    "pending_sell_level",
]
TREND_COLUMNS = [
    "date",
    "symbol",
    "name",
    "reference_price",
    "boll_upper",
    "boll_mid",
    "boll_lower",
    "atr20",
    "trend_buy_trigger",
    "trend_reduce_trigger",
    "trend_exit_trigger",
    "trend_batches",
    "max_position_pct",
    "trigger_status",
    "filled_status",
]
GRID_COLUMNS = [
    "date",
    "symbol",
    "name",
    "reference_price",
    "grid_lower",
    "grid_mid",
    "grid_upper",
    "grid_max_layers",
    "grid_buy_levels",
    "grid_sell_levels",
    "grid_total_max_position_pct",
    "max_position_pct",
    "trigger_status",
    "filled_status",
]
PENDING_COLUMNS = [
    "date",
    "symbol",
    "name",
    "stock_mode",
    "available_shares",
    "today_bought_shares",
    "pending_sell_level",
    "trigger_status",
    "filled_status",
]
RISK_COLUMNS = ["date", "symbol", "name", "stock_mode", "trigger_status", "filled_status", "failed_reason", "risk_note", "reason"]


@dataclass(frozen=True)
class T1ThermostatReport:
    tables: dict[str, pd.DataFrame]
    report_date: str


def build_t1_thermostat_report(trigger_plan: pd.DataFrame | None, errors: pd.DataFrame | None) -> T1ThermostatReport:
    plan = _copy_frame(trigger_plan)
    error_frame = _copy_frame(errors)
    report_date = _report_date(plan)
    tables = {
        "市场状态与仓位折扣": _localize(_select(plan.drop_duplicates(subset=_existing(plan, MARKET_COLUMNS)), MARKET_COLUMNS)),
        "个股模式摘要": _localize(_select(plan, SUMMARY_COLUMNS)),
        "趋势触发计划": _localize(_select(_filter_equals(plan, "stock_mode", "trend"), TREND_COLUMNS)),
        "网格触发计划": _localize(_select(_filter_equals(plan, "stock_mode", "range"), GRID_COLUMNS)),
        "待卖记录": _localize(_select(_filter_nonempty(plan, "pending_sell_level"), PENDING_COLUMNS)),
        "失败原因和风险提示": _localize(_select(_risk_rows(plan), RISK_COLUMNS)),
        "错误与数据质量": _localize(error_frame),
        "详细字段": _localize(plan),
    }
    return T1ThermostatReport(tables=tables, report_date=report_date)


def default_t1_thermostat_report_filename(report: T1ThermostatReport) -> str:
    return f"t1_thermostat_report_{report.report_date}.xlsx"


def export_t1_thermostat_excel(report: T1ThermostatReport, output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in EXPECTED_T1_THERMOSTAT_SHEETS:
            frame = report.tables.get(sheet_name, pd.DataFrame())
            _write_sheet(writer, sheet_name, frame)
    return path


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    safe = sheet_name[:31]
    if frame is None or frame.empty:
        pd.DataFrame([["暂无数据"]]).to_excel(writer, sheet_name=safe, index=False, header=False)
    else:
        frame.to_excel(writer, sheet_name=safe, index=False)
    worksheet = writer.sheets[safe]
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    if frame is not None and not frame.empty and frame.shape[1] > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True if _is_wrapped_header(worksheet, cell.column) else cell.alignment.wrap_text)
            _apply_number_format(worksheet, cell)
    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(max((len(value) for value in values), default=8) + 2, 10), 48)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _is_wrapped_header(worksheet, column_index: int) -> bool:
    header = worksheet.cell(row=1, column=column_index).value
    raw = _raw_column_name(str(header or ""))
    return raw in TEXT_COLUMNS


def _apply_number_format(worksheet, cell) -> None:
    if cell.row == 1:
        return
    raw = _raw_column_name(str(worksheet.cell(row=1, column=cell.column).value or ""))
    if raw in PERCENT_COLUMNS:
        cell.number_format = "0.00%"
    elif raw in INTEGER_COLUMNS:
        cell.number_format = "0"
    elif raw in DECIMAL_COLUMNS:
        cell.number_format = "0.00"


def _raw_column_name(label: str) -> str:
    reverse = {value: key for key, value in COLUMN_LABELS.items()}
    return reverse.get(label, label)


def _localize(frame: pd.DataFrame) -> pd.DataFrame:
    data = _copy_frame(frame)
    return data.rename(columns={column: COLUMN_LABELS.get(column, column) for column in data.columns})


def _select(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame(columns=columns)
    return frame.reindex(columns=[column for column in columns if column in frame.columns])


def _filter_equals(frame: pd.DataFrame, column: str, value: str) -> pd.DataFrame:
    if frame is None or frame.empty or column not in frame:
        return pd.DataFrame()
    return frame[frame[column].fillna("").astype(str) == value].copy()


def _filter_nonempty(frame: pd.DataFrame, column: str) -> pd.DataFrame:
    if frame is None or frame.empty or column not in frame:
        return pd.DataFrame()
    return frame[frame[column].fillna("").astype(str).str.strip().ne("")].copy()


def _risk_rows(frame: pd.DataFrame) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame()
    failed = frame["failed_reason"].fillna("").astype(str).str.strip().ne("") if "failed_reason" in frame else pd.Series(False, index=frame.index)
    risk = frame["risk_note"].fillna("").astype(str).str.strip().ne("") if "risk_note" in frame else pd.Series(False, index=frame.index)
    return frame[failed | risk].copy()


def _existing(frame: pd.DataFrame, columns: list[str]) -> list[str]:
    return [column for column in columns if frame is not None and column in frame.columns]


def _copy_frame(frame: pd.DataFrame | None) -> pd.DataFrame:
    if frame is None:
        return pd.DataFrame()
    return frame.copy()


def _report_date(trigger_plan: pd.DataFrame) -> str:
    if trigger_plan is not None and not trigger_plan.empty and "date" in trigger_plan:
        values = trigger_plan["date"].dropna().astype(str).str.replace("-", "", regex=False)
        values = values[values.str.len() >= 8]
        if not values.empty:
            return values.iloc[0][:8]
    return datetime.now().strftime("%Y%m%d")

