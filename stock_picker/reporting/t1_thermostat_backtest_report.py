from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from threading import Lock
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import Alignment

if TYPE_CHECKING:
    from stock_picker.strategies.thermostat_backtest import T1ThermostatBacktestResult


EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS = [
    "回测摘要",
    "回测说明",
    "参数与账户设置",
    "数据来源与股票池",
    "每日资产",
    "权益与回撤",
    "每日持仓",
    "每日触发计划",
    "订单明细",
    "成交明细",
    "失败订单",
    "取消订单",
    "pending明细",
    "趋势批次",
    "网格层级",
    "个股表现",
    "趋势策略表现",
    "网格策略表现",
    "市场状态表现",
    "数据质量",
    "公司行为影响",
]


COLUMN_LABELS = {
    "initial_asset": "初始资产",
    "final_asset": "期末资产",
    "total_return": "总收益率",
    "annualized_return": "年化收益率",
    "benchmark_return": "基准收益率",
    "excess_return": "超额收益率",
    "max_drawdown": "最大回撤",
    "sharpe_ratio": "夏普比率",
    "annual_volatility": "年化波动率",
    "completed_cycle_count": "已完成交易周期数",
    "completed_cycle_win_rate": "已完成周期胜率",
    "profit_loss_ratio": "盈亏比",
    "average_win": "平均盈利",
    "average_loss": "平均亏损",
    "average_holding_days": "平均持有天数",
    "trade_count": "成交笔数",
    "buy_count": "买入笔数",
    "sell_count": "卖出笔数",
    "failed_order_count": "失败订单数",
    "pending_order_count": "pending订单数",
    "pending_average_duration_days": "pending平均持续天数",
    "average_position_utilization": "平均仓位利用率",
    "max_position_utilization": "最大仓位利用率",
    "average_cash_ratio": "平均现金比例",
    "missing_data_ratio": "缺失数据比例",
    "ambiguity_count": "日线歧义次数",
    "corporate_action_affected_symbol_count": "公司行为影响股票数",
    "corporate_action_affected_date_count": "公司行为影响日期数",
    "trading_day_count": "交易日数",
    "lifecycle_order_detail_count": "订单明细数",
    "fill_detail_count": "成交明细数",
    "failed_cancelled_detail_count": "失败或撤单数",
    "pending_detail_count": "pending记录数",
    "trend_batch_detail_count": "趋势批次数",
    "grid_layer_detail_count": "网格层级数",
    "data_quality_detail_count": "数据质量记录数",
    "corporate_action_detail_count": "公司行为记录数",
    "date": "日期",
    "trade_date": "交易日期",
    "start_date": "开始日期",
    "end_date": "结束日期",
    "buy_date": "买入日期",
    "sell_date": "卖出日期",
    "fill_date": "成交日期",
    "first_fill_date": "首次成交日期",
    "last_fill_date": "末次成交日期",
    "pending_since": "pending起始日期",
    "last_attempt_date": "最近尝试日期",
    "data_cutoff_date": "数据截止日期",
    "symbol": "股票代码",
    "code": "代码",
    "name": "股票名称",
    "cash": "现金",
    "position_value": "持仓市值",
    "total_asset": "总资产",
    "cash_ratio": "现金比例",
    "position_ratio": "持仓比例",
    "realized_pnl": "已实现盈亏",
    "unrealized_pnl": "未实现盈亏",
    "daily_return": "日收益率",
    "cumulative_return": "累计收益率",
    "running_peak": "历史权益峰值",
    "drawdown": "回撤",
    "precision": "回测精度",
    "precision_disclosure": "精度说明",
    "approximate_intraday_sequence": "盘中顺序为近似",
    "approximation_warnings": "近似处理警告",
    "stock_mode": "股票模式",
    "mode": "股票模式",
    "family": "策略族",
    "market_regime": "市场状态",
    "market_regime_normalized": "标准化市场状态",
    "market_position_discount": "市场仓位折扣",
    "stock_regime": "个股状态",
    "reference_price": "参考价格",
    "boll_upper": "布林上轨",
    "boll_mid": "布林中轨",
    "boll_lower": "布林下轨",
    "atr20": "ATR20",
    "volume_ma20": "20日平均成交量",
    "trend_buy_trigger": "趋势买入触发价",
    "trend_reduce_trigger": "趋势减仓触发价",
    "trend_exit_trigger": "趋势退出触发价",
    "effective_trend_exit_trigger": "有效趋势退出触发价",
    "trend_batches": "趋势分批",
    "grid_lower": "网格下沿",
    "grid_mid": "网格中枢",
    "grid_upper": "网格上沿",
    "grid_max_layers": "最大网格层数",
    "configured_grid_layers": "配置网格层数",
    "effective_grid_layers": "有效网格层数",
    "grid_layer_spacing_pct": "网格层间距",
    "grid_buy_levels": "网格买入层",
    "grid_sell_levels": "网格卖出层",
    "grid_total_max_position_pct": "网格总仓位上限",
    "target_position_pct": "目标仓位比例",
    "max_position_pct": "单股仓位上限",
    "total_shares": "总股数",
    "available_shares": "可卖股数",
    "today_bought_shares": "当日买入股数",
    "trend_shares": "趋势持仓股数",
    "grid_shares": "网格持仓股数",
    "average_cost": "平均成本",
    "close": "收盘价",
    "market_value": "市值",
    "pending_sell_level": "pending卖出级别",
    "pending_count": "pending数量",
    "share_split_source": "股数拆分来源",
    "trigger_status": "触发状态",
    "filled_status": "成交状态",
    "failed_reason": "失败原因",
    "risk_note": "风险提示",
    "reason": "原因",
    "data_sufficient": "数据是否充足",
    "order_id": "订单ID",
    "trigger_type": "触发类型",
    "side": "买卖方向",
    "status": "状态",
    "trigger_price": "触发价格",
    "base_price": "基准价格",
    "execution_price": "成交价格",
    "intended_shares": "计划股数",
    "actual_shares": "实际股数",
    "gross_amount": "成交金额",
    "commission": "佣金",
    "stamp_tax": "印花税",
    "slippage_cost": "滑点成本",
    "total_cost": "总成本",
    "cash_before": "成交前现金",
    "cash_after": "成交后现金",
    "position_before": "成交前持仓股数",
    "position_after": "成交后持仓股数",
    "pending_level": "pending级别",
    "grid_layer": "网格层ID",
    "trend_batch": "趋势批次",
    "risk_rank": "风险排序",
    "plan_priority": "计划优先级",
    "quality_warning": "质量警告",
    "failure_reason": "失败原因",
    "plan_trace_id": "计划追踪ID",
    "candidate_trace_id": "候选追踪ID",
    "episode_id": "pending事件ID",
    "event_type": "事件类型",
    "is_terminal": "是否终态",
    "level": "级别",
    "owner_id": "所有者ID",
    "remaining_shares": "剩余股数",
    "requested_shares": "请求股数",
    "duration_days": "持续交易日数",
    "attempt_count": "尝试次数",
    "last_failure": "最近失败原因",
    "source_order_id": "来源订单ID",
    "batch_index": "批次序号",
    "target_ratio": "目标比例",
    "planned_shares": "计划股数",
    "filled_shares": "已成交股数",
    "fill_price": "成交价格",
    "layer_id": "层级ID",
    "buy_price": "买入价格",
    "sell_price": "卖出价格",
    "target_shares": "目标股数",
    "held_shares": "持有股数",
    "buy_cost": "买入成本",
    "key": "统计对象",
    "total_pnl": "总盈亏",
    "closed_cost": "已平仓成本",
    "open_cost": "未平仓成本",
    "invested_cost": "投入成本",
    "return": "收益率",
    "completed_cycles": "已完成周期数",
    "wins": "盈利周期数",
    "win_rate": "胜率",
    "benchmark_symbol": "基准代码",
    "trading_days": "交易日数",
    "annualized_benchmark_return": "基准年化收益率",
    "severity": "严重程度",
    "stream": "数据流",
    "message": "说明",
    "details": "详情",
    "observation_expected": "应有观测数",
    "observation_missing": "缺失观测数",
    "evidence": "证据",
    "parameter_name": "参数名",
    "parameter_value": "参数值",
    "parameter_source": "参数来源",
    "user_overridden": "是否用户覆盖",
    "note": "备注",
    "metadata_key": "元数据项",
    "metadata_value": "元数据值",
    "cycle_id": "交易周期ID",
    "buy_order_id": "买入订单ID",
    "sell_order_id": "卖出订单ID",
    "shares": "股数",
    "buy_fees": "买入费用",
    "sell_fees": "卖出费用",
    "gross_pnl": "毛盈亏",
    "net_pnl": "净盈亏",
    "return_pct": "收益率",
    "holding_days": "持有天数",
    "is_win": "是否盈利",
    "record_type": "记录类型",
}

PERCENT_COLUMNS = {
    "total_return", "annualized_return", "benchmark_return", "excess_return",
    "max_drawdown", "annual_volatility", "completed_cycle_win_rate",
    "average_position_utilization", "max_position_utilization",
    "average_cash_ratio", "missing_data_ratio", "cash_ratio", "position_ratio",
    "daily_return", "cumulative_return", "drawdown", "market_position_discount",
    "grid_layer_spacing_pct", "grid_total_max_position_pct", "target_position_pct",
    "max_position_pct", "target_ratio", "return", "win_rate",
    "annualized_benchmark_return", "return_pct",
}
MONEY_COLUMNS = {
    "initial_asset", "final_asset", "average_win", "average_loss", "cash",
    "position_value", "total_asset", "running_peak", "realized_pnl",
    "unrealized_pnl", "market_value", "gross_amount", "commission",
    "stamp_tax", "slippage_cost", "total_cost", "cash_before", "cash_after",
    "buy_cost", "total_pnl", "closed_cost", "open_cost", "invested_cost",
    "buy_fees", "sell_fees", "gross_pnl", "net_pnl",
}
PRICE_COLUMNS = {
    "reference_price", "boll_upper", "boll_mid", "boll_lower", "atr20",
    "trend_buy_trigger", "trend_reduce_trigger", "trend_exit_trigger",
    "effective_trend_exit_trigger", "grid_lower", "grid_mid", "grid_upper",
    "average_cost", "close", "trigger_price", "base_price", "execution_price",
    "fill_price", "buy_price", "sell_price",
}
INTEGER_COLUMNS = {
    "completed_cycle_count", "trade_count", "buy_count", "sell_count",
    "failed_order_count", "pending_order_count", "ambiguity_count",
    "corporate_action_affected_symbol_count", "corporate_action_affected_date_count",
    "trading_day_count", "lifecycle_order_detail_count", "fill_detail_count",
    "failed_cancelled_detail_count", "pending_detail_count",
    "trend_batch_detail_count", "grid_layer_detail_count", "data_quality_detail_count",
    "corporate_action_detail_count", "total_shares", "available_shares",
    "today_bought_shares", "trend_shares", "grid_shares", "pending_count",
    "grid_max_layers", "configured_grid_layers", "effective_grid_layers",
    "intended_shares", "actual_shares", "position_before", "position_after",
    "risk_rank", "plan_priority", "remaining_shares", "requested_shares",
    "duration_days", "attempt_count", "batch_index", "planned_shares",
    "filled_shares", "target_shares", "held_shares", "completed_cycles", "wins",
    "trading_days", "observation_expected", "observation_missing", "shares",
    "holding_days",
}
DATE_COLUMNS = {
    "date", "trade_date", "start_date", "end_date", "buy_date", "sell_date",
    "fill_date", "first_fill_date", "last_fill_date", "pending_since",
    "last_attempt_date", "data_cutoff_date",
}
ID_COLUMNS = {
    "symbol", "benchmark_symbol", "order_id", "plan_trace_id",
    "candidate_trace_id", "episode_id", "owner_id", "source_order_id",
    "layer_id", "grid_layer", "cycle_id", "buy_order_id", "sell_order_id",
}
LONG_TEXT_COLUMNS = {
    "precision_disclosure", "approximation_warnings", "grid_buy_levels",
    "grid_sell_levels", "failed_reason", "risk_note", "reason", "quality_warning",
    "failure_reason", "last_failure", "message", "details", "evidence", "note",
    "metadata_value", "parameter_value",
}

_DEFAULT_PARAMETER_DISCLOSURES = [
    ("trend_symbol_base_max", 0.20, "系统默认"),
    ("trend_total_base_max", 0.65, "系统默认"),
    ("grid_symbol_base_max", 0.15, "系统默认"),
    ("grid_total_hard_max", 0.40, "系统默认"),
    ("account_total_max", 0.95, "系统默认"),
    ("force_final_liquidation", False, "系统默认"),
]

_TIMESTAMP_LOCK = Lock()
_LAST_DEFAULT_TIMESTAMP: datetime | None = None


@dataclass(frozen=True)
class T1ThermostatBacktestReport:
    tables: dict[str, pd.DataFrame]
    generated_at: datetime


def build_t1_thermostat_backtest_report(
    result: T1ThermostatBacktestResult,
    *,
    generated_at: datetime | None = None,
) -> T1ThermostatBacktestReport:
    raw = {name: frame.copy(deep=True) for name, frame in result.tables.items()}
    failed_cancelled = raw["failed_cancelled_orders"]
    failed = _status_rows(failed_cancelled, {"failed", "expired"})
    cancelled = _status_rows(failed_cancelled, {"cancelled"})
    pending = _enrich_pending_trace_ids(raw["pending_history"], raw["lifecycle_orders"])
    fills_and_cycles = _fills_and_cycles(raw["fills"], raw["closed_trade_cycles"])

    tables = {
        "回测摘要": _localize(_summary(raw)),
        "回测说明": _explanation_table(),
        "参数与账户设置": _localize(_parameter_table(raw)),
        "数据来源与股票池": _localize(_source_and_pool_table(raw)),
        "每日资产": _localize(raw["daily_assets"]),
        "权益与回撤": _localize(raw["equity_drawdown"]),
        "每日持仓": _localize(raw["daily_positions"]),
        "每日触发计划": _localize(raw["daily_trigger_plans"]),
        "订单明细": _localize(raw["lifecycle_orders"]),
        "成交明细": _localize(fills_and_cycles),
        "失败订单": _localize(failed),
        "取消订单": _localize(cancelled),
        "pending明细": _localize(pending),
        "趋势批次": _localize(raw["trend_batches"]),
        "网格层级": _localize(raw["grid_layers"]),
        "个股表现": _localize(raw["symbol_performance"]),
        "趋势策略表现": _localize(raw["trend_performance"]),
        "网格策略表现": _localize(raw["grid_performance"]),
        "市场状态表现": _localize(raw["market_performance"]),
        "数据质量": _localize(raw["data_quality"]),
        "公司行为影响": _localize(raw["corporate_actions"]),
    }
    return T1ThermostatBacktestReport(
        tables=tables,
        generated_at=generated_at or _next_default_timestamp(),
    )


def default_t1_thermostat_backtest_report_filename(
    value: T1ThermostatBacktestReport | datetime | None = None,
) -> str:
    if isinstance(value, T1ThermostatBacktestReport):
        timestamp = value.generated_at
    elif isinstance(value, datetime):
        timestamp = value
    else:
        timestamp = datetime.now()
    return f"t1_thermostat_backtest_{timestamp:%Y%m%d_%H%M%S}.xlsx"


def export_t1_thermostat_backtest_excel(
    report: T1ThermostatBacktestReport,
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in EXPECTED_T1_THERMOSTAT_BACKTEST_SHEETS:
            frame = report.tables.get(sheet_name, pd.DataFrame()).copy(deep=True)
            _write_sheet(writer, sheet_name, frame)
    return path


def _summary(raw: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frame = raw["summary"].copy(deep=True)
    if frame.empty:
        frame = pd.DataFrame([{}])
    counts = {
        "lifecycle_order_detail_count": len(raw["lifecycle_orders"]),
        "fill_detail_count": len(raw["fills"]),
        "failed_cancelled_detail_count": len(raw["failed_cancelled_orders"]),
        "pending_detail_count": len(raw["pending_history"]),
        "trend_batch_detail_count": len(raw["trend_batches"]),
        "grid_layer_detail_count": len(raw["grid_layers"]),
        "data_quality_detail_count": len(raw["data_quality"]),
        "corporate_action_detail_count": len(raw["corporate_actions"]),
    }
    for name, value in counts.items():
        frame[name] = value
    return frame


def _next_default_timestamp() -> datetime:
    global _LAST_DEFAULT_TIMESTAMP
    candidate = datetime.now().replace(microsecond=0)
    with _TIMESTAMP_LOCK:
        if _LAST_DEFAULT_TIMESTAMP is not None and candidate <= _LAST_DEFAULT_TIMESTAMP:
            candidate = _LAST_DEFAULT_TIMESTAMP + timedelta(seconds=1)
        _LAST_DEFAULT_TIMESTAMP = candidate
    return candidate


def _explanation_table() -> pd.DataFrame:
    rows = [
        ("回测精度", "回测精度：日线近似"),
        ("分钟线", "分钟线：未使用"),
        ("盘中时点", "盘中触发时间：无法准确识别"),
        ("同日多触发", "同日多触发：使用保守顺序处理"),
        ("触发顺序", "pending卖出、风险控制卖出、趋势退出、趋势减仓、网格卖出、趋势买入、网格买入。"),
        ("趋势买入基准价", "趋势买入成交基准价取触发价与当日收盘价的较高值。"),
        ("趋势卖出基准价", "趋势卖出成交基准价取触发价与当日收盘价的较低值。"),
        ("网格基准价", "网格买卖以对应网格层价格作为成交基准价。"),
        ("pending卖出基准价", "pending卖出以下一交易日开盘价作为成交基准价。"),
        ("买入成交价", "买入成交价 = 基准价 ×（1 + 滑点率），并受涨停价上限约束。"),
        ("卖出成交价", "卖出成交价 = 基准价 ×（1 - 滑点率），并受跌停价下限约束。"),
        ("涨停处理", "一字涨停或无法买入时保守记录失败，不虚构成交。"),
        ("跌停处理", "跌停无法卖出时保留pending状态并在后续交易日重试。"),
        ("停牌处理", "停牌或无有效价格时不成交，保留失败或pending生命周期记录。"),
        ("T+1约束", "当日买入股数不可当日卖出；触及卖出条件时转为pending。"),
        (
            "追踪标识语义",
            "未创建执行候选时，candidate_trace_id按设计留空；订单ID与计划追踪ID仍原样保留。",
        ),
        ("期末处理", "force_final_liquidation=False，期末持仓按收盘价估值，不强制平仓。"),
    ]
    return pd.DataFrame(rows, columns=["项目", "说明"])


def _parameter_table(raw: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frame = raw["parameters"].copy(deep=True)
    required_columns = [
        "parameter_name", "parameter_value", "parameter_source",
        "user_overridden", "note",
    ]
    frame = frame.reindex(columns=required_columns)
    present = set(frame["parameter_name"].dropna().astype(str))
    additions: list[dict[str, object]] = []
    for name, value, source in _DEFAULT_PARAMETER_DISCLOSURES:
        if name not in present:
            additions.append(_parameter_row(name, value, source))
    settings = _parameter_values(frame)
    if "趋势分批比例" not in present:
        additions.append(_parameter_row("趋势分批比例", "40%, 35%, 25%", "策略固定规则"))
    layer_count = _first_non_null(
        raw["daily_trigger_plans"], ["configured_grid_layers", "grid_max_layers"], 3,
    )
    additions.append(_parameter_row("网格层数", layer_count, "触发计划/策略规则"))
    spacing = _first_non_null(raw["daily_trigger_plans"], ["grid_layer_spacing_pct"], "按波动率动态计算")
    additions.append(_parameter_row("网格间距", spacing, "触发计划/策略规则"))
    additions.extend([
        _parameter_row(
            "买入成交价公式",
            "min(基准价 × (1 + slippage_pct), 涨停价)",
            "保守成交规则",
        ),
        _parameter_row(
            "卖出成交价公式",
            "max(基准价 × (1 - slippage_pct), 跌停价)",
            "保守成交规则",
        ),
        _parameter_row(
            "买入费用公式",
            "max(成交金额 × commission_rate, minimum_commission)",
            "账户费用规则",
        ),
        _parameter_row(
            "卖出费用公式",
            "max(成交金额 × commission_rate, minimum_commission) + 成交金额 × stamp_tax_rate",
            "账户费用规则",
        ),
    ])
    if additions:
        frame = pd.concat([frame, pd.DataFrame(additions)], ignore_index=True)
    # Preserve actual values from the runner; this assignment only documents their
    # relationship and never substitutes report defaults for present settings.
    frame["note"] = frame["note"].fillna("")
    if settings:
        frame.loc[frame["parameter_name"].eq("force_final_liquidation"), "note"] = "必须为False"
    return frame


def _source_and_pool_table(raw: dict[str, pd.DataFrame]) -> pd.DataFrame:
    parameters = _parameter_values(raw["parameters"])
    refresh = parameters.get("refresh", False)
    quality = raw["data_quality"]
    quality_codes = quality.get("code", pd.Series(dtype=str)).fillna("").astype(str)
    rows = [
        ("数据源", parameters.get("source", "由回测请求与本地缓存确定")),
        ("指标数据流", "前复权指标流（qfq），仅使用交易日前历史生成计划"),
        ("执行数据流", "不复权执行流（bfq），用于触发、成交与收盘估值"),
        ("缓存策略", "强制刷新后更新缓存" if _truthy(refresh) else "优先读取本地缓存，按缺口补取"),
        ("缓存缺口记录数", int(quality_codes.str.contains("cache_gap", case=False).sum())),
        ("预热不足记录数", int(quality_codes.str.contains("warmup", case=False).sum())),
    ]
    metadata = raw["stock_pool_metadata"]
    for row in metadata.to_dict("records"):
        rows.append((row.get("metadata_key", ""), row.get("metadata_value", "")))
    keys = {str(key) for key, _ in rows}
    defaults = [
        ("membership", "未提供；按提交时静态股票池解释"),
        ("generation_method", "未提供"),
        ("look_ahead_selection_warning", "未提供；使用事后筛选股票池可能产生前视选择偏差"),
        ("survivor_bias_warning", "未提供；当前成分股票池可能产生幸存者偏差"),
    ]
    rows.extend(item for item in defaults if item[0] not in keys)
    return pd.DataFrame(rows, columns=["metadata_key", "metadata_value"])


def _status_rows(frame: pd.DataFrame, statuses: set[str]) -> pd.DataFrame:
    if frame.empty or "status" not in frame:
        return frame.iloc[0:0].copy()
    mask = frame["status"].fillna("").astype(str).str.lower().isin(statuses)
    return frame.loc[mask].copy()


def _enrich_pending_trace_ids(pending: pd.DataFrame, orders: pd.DataFrame) -> pd.DataFrame:
    data = pending.copy(deep=True)
    if data.empty:
        return data
    if "source_order_id" not in data or "order_id" not in orders:
        return data
    lookup_columns = [
        column for column in ("order_id", "candidate_trace_id") if column in orders
    ]
    if len(lookup_columns) < 2:
        return data
    lookup = orders[lookup_columns].drop_duplicates("order_id", keep="last")
    data = data.merge(
        lookup,
        how="left",
        left_on="source_order_id",
        right_on="order_id",
        suffixes=("", "_source"),
    )
    if "order_id" in data:
        data = data.drop(columns=["order_id"])
    return data


def _fills_and_cycles(fills: pd.DataFrame, cycles: pd.DataFrame) -> pd.DataFrame:
    fill_rows = fills.copy(deep=True)
    cycle_rows = cycles.copy(deep=True)
    if not fill_rows.empty:
        fill_rows.insert(0, "record_type", "成交订单")
    if not cycle_rows.empty:
        cycle_rows.insert(0, "record_type", "已完成交易周期")
    if fill_rows.empty:
        return cycle_rows
    if cycle_rows.empty:
        return fill_rows
    return pd.concat([fill_rows, cycle_rows], ignore_index=True, sort=False)


def _parameter_row(name: str, value: object, source: str) -> dict[str, object]:
    return {
        "parameter_name": name,
        "parameter_value": value,
        "parameter_source": source,
        "user_overridden": False,
        "note": "",
    }


def _parameter_values(frame: pd.DataFrame) -> dict[str, object]:
    if frame.empty or not {"parameter_name", "parameter_value"}.issubset(frame.columns):
        return {}
    return {
        str(row["parameter_name"]): row["parameter_value"]
        for row in frame[["parameter_name", "parameter_value"]].to_dict("records")
    }


def _first_non_null(frame: pd.DataFrame, columns: list[str], default: object) -> object:
    for column in columns:
        if column in frame:
            values = frame[column].dropna()
            if not values.empty:
                return values.iloc[0]
    return default


def _truthy(value: object) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def _localize(frame: pd.DataFrame) -> pd.DataFrame:
    data = frame.copy(deep=True)
    return data.rename(columns={
        column: COLUMN_LABELS.get(str(column), f"扩展字段（{column}）")
        for column in data.columns
    })


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    safe = sheet_name[:31]
    if frame.empty:
        pd.DataFrame([["暂无数据"]]).to_excel(
            writer, sheet_name=safe, index=False, header=False,
        )
    else:
        _excel_values(frame).to_excel(writer, sheet_name=safe, index=False)
    worksheet = writer.sheets[safe]
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    if not frame.empty and frame.shape[1] > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    reverse_labels = {label: raw for raw, label in COLUMN_LABELS.items()}
    for row in worksheet.iter_rows():
        for cell in row:
            raw_column = reverse_labels.get(str(worksheet.cell(1, cell.column).value or ""), "")
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(raw_column in LONG_TEXT_COLUMNS or _needs_wrap(cell.value)),
            )
            if cell.row > 1:
                _apply_number_format(cell, raw_column, worksheet)
    for column_cells in worksheet.columns:
        display_lengths = [_display_width(cell.value) for cell in column_cells]
        width = min(max(max(display_lengths, default=8) + 2, 10), 48)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _excel_values(frame: pd.DataFrame) -> pd.DataFrame:
    data = frame.copy(deep=True)
    reverse_labels = {label: raw for raw, label in COLUMN_LABELS.items()}
    for column in data.columns:
        raw = reverse_labels.get(str(column), "")
        if raw in DATE_COLUMNS:
            data[column] = pd.to_datetime(data[column], errors="coerce")
        elif raw in ID_COLUMNS:
            data[column] = data[column].map(
                lambda value: "" if pd.isna(value) else str(value)
            )
    return data


def _apply_number_format(cell, raw_column: str, worksheet) -> None:
    if raw_column in PERCENT_COLUMNS:
        cell.number_format = "0.00%"
    elif raw_column in MONEY_COLUMNS:
        cell.number_format = "#,##0.00"
    elif raw_column in PRICE_COLUMNS:
        cell.number_format = "0.000"
    elif raw_column in INTEGER_COLUMNS:
        cell.number_format = "0"
    elif raw_column in DATE_COLUMNS:
        cell.number_format = "yyyy-mm-dd"
    elif raw_column in ID_COLUMNS:
        cell.number_format = "@"
    elif raw_column == "parameter_value":
        parameter_name = worksheet.cell(cell.row, 1).value
        if parameter_name in {
            "commission_rate", "stamp_tax_rate", "slippage_pct",
            "trend_symbol_base_max", "trend_total_base_max", "grid_symbol_base_max",
            "grid_total_hard_max", "account_total_max", "网格间距",
        }:
            cell.number_format = "0.00%"


def _needs_wrap(value: object) -> bool:
    if value is None:
        return False
    text = str(value)
    return len(text) > 24 or "\n" in text


def _display_width(value: object) -> int:
    if value is None:
        return 0
    return sum(2 if ord(character) > 127 else 1 for character in str(value))
